from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse

from invoices.models import (
    Counterparty,
    Invoice,
    OCRJob,
    ResponsiblePerson,
)


class InvoiceBotReportDetailTests(TestCase):

    def setUp(self):
        User = get_user_model()

        self.user = User.objects.create_user(
            username="bot-report-detail-user",
            email="bot-report-detail-user@example.com",
            password="pass12345",
            is_staff=True,
        )

        self.counterparty = Counterparty.objects.create(
            name="ООО ДЕТАЛИЗАЦИЯ БОТА",
            full_name="Общество с ограниченной ответственностью ДЕТАЛИЗАЦИЯ БОТА",
            inn="7701234567",
            kpp="770101001",
            bank_name="АО ТЕСТ БАНК",
            account_number="40702810000000000001",
            bik="044525225",
            source=Counterparty.SOURCE_1C,
            is_active=True,
        )

        self.responsible = ResponsiblePerson.objects.create(
            full_name="Ответственный детализации бота",
            is_active=True,
        )

    def create_invoice(self, **kwargs):
        defaults = {
            "user": self.user,
            "title": "BOT DETAIL INVOICE",
            "amount": Decimal("1000.00"),
            "amount_verified": True,
            "status": Invoice.STATUS_APPROVED,
            "document_type": Invoice.DOCUMENT_TYPE_INVOICE,
            "document_date": date(2026, 7, 1),
            "planned_payment_date": date(2026, 7, 10),
            "responsible": self.responsible,
            "counterparty": self.counterparty,
            "ocr_text": "OCR TEXT",
        }
        defaults.update(
            kwargs
        )

        return Invoice.objects.create(
            **defaults
        )

    def test_ready_category_shows_only_ready_invoices(self):
        ready_invoice = self.create_invoice(
            title="READY DETAIL INVOICE",
        )
        not_ready_invoice = self.create_invoice(
            title="NOT READY DETAIL INVOICE",
            planned_payment_date=None,
        )

        self.client.force_login(
            self.user
        )

        response = self.client.get(
            reverse(
                "invoice_bot_report_detail",
                kwargs={
                    "category": "ready",
                },
            )
        )

        self.assertEqual(
            response.status_code,
            200,
        )
        self.assertContains(
            response,
            "Готовы к реестру",
        )
        self.assertContains(
            response,
            ready_invoice.title,
        )
        self.assertNotContains(
            response,
            not_ready_invoice.title,
        )
        self.assertContains(
            response,
            "Нет блокировок",
        )

    def test_not_ready_category_shows_blocked_invoices(self):
        ready_invoice = self.create_invoice(
            title="READY DETAIL INVOICE",
        )
        not_ready_invoice = self.create_invoice(
            title="NO DATE DETAIL INVOICE",
            planned_payment_date=None,
        )

        self.client.force_login(
            self.user
        )

        response = self.client.get(
            reverse(
                "invoice_bot_report_detail",
                kwargs={
                    "category": "not-ready",
                },
            )
        )

        self.assertEqual(
            response.status_code,
            200,
        )
        self.assertContains(
            response,
            "Не готовы к реестру",
        )
        self.assertContains(
            response,
            not_ready_invoice.title,
        )
        self.assertNotContains(
            response,
            ready_invoice.title,
        )
        self.assertContains(
            response,
            "Причины / проверка",
        )

    def test_without_planned_payment_date_category_filters_invoices(self):
        matching_invoice = self.create_invoice(
            title="WITHOUT DATE DETAIL INVOICE",
            planned_payment_date=None,
        )
        other_invoice = self.create_invoice(
            title="WITH DATE DETAIL INVOICE",
        )

        self.client.force_login(
            self.user
        )

        response = self.client.get(
            reverse(
                "invoice_bot_report_detail",
                kwargs={
                    "category": "without-planned-payment-date",
                },
            )
        )

        self.assertContains(
            response,
            matching_invoice.title,
        )
        self.assertNotContains(
            response,
            other_invoice.title,
        )

    def test_without_counterparty_category_filters_invoices(self):
        matching_invoice = self.create_invoice(
            title="WITHOUT COUNTERPARTY DETAIL INVOICE",
            counterparty=None,
        )
        other_invoice = self.create_invoice(
            title="WITH COUNTERPARTY DETAIL INVOICE",
        )

        self.client.force_login(
            self.user
        )

        response = self.client.get(
            reverse(
                "invoice_bot_report_detail",
                kwargs={
                    "category": "without-counterparty",
                },
            )
        )

        self.assertContains(
            response,
            matching_invoice.title,
        )
        self.assertNotContains(
            response,
            other_invoice.title,
        )

    def test_unverified_amount_category_filters_invoices(self):
        matching_invoice = self.create_invoice(
            title="AMOUNT NEEDS CHECK DETAIL INVOICE",
            amount_verified=False,
        )
        other_invoice = self.create_invoice(
            title="AMOUNT OK DETAIL INVOICE",
        )

        self.client.force_login(
            self.user
        )

        response = self.client.get(
            reverse(
                "invoice_bot_report_detail",
                kwargs={
                    "category": "unverified-amount",
                },
            )
        )

        self.assertContains(
            response,
            matching_invoice.title,
        )
        self.assertNotContains(
            response,
            other_invoice.title,
        )

    def test_without_ocr_text_category_filters_invoices(self):
        matching_invoice = self.create_invoice(
            title="WITHOUT OCR DETAIL INVOICE",
            ocr_text="",
        )
        other_invoice = self.create_invoice(
            title="WITH OCR DETAIL INVOICE",
        )

        self.client.force_login(
            self.user
        )

        response = self.client.get(
            reverse(
                "invoice_bot_report_detail",
                kwargs={
                    "category": "without-ocr-text",
                },
            )
        )

        self.assertContains(
            response,
            matching_invoice.title,
        )
        self.assertNotContains(
            response,
            other_invoice.title,
        )

    def test_unknown_document_type_category_filters_invoices(self):
        matching_invoice = self.create_invoice(
            title="MYSTERY DOC TYPE DETAIL INVOICE",
            document_type=Invoice.DOCUMENT_TYPE_UNKNOWN,
            ocr_text="Акт сверки взаимных расчетов за июль 2026",
        )
        without_ocr_invoice = self.create_invoice(
            title="NO OCR MYSTERY TYPE DETAIL INVOICE",
            document_type=Invoice.DOCUMENT_TYPE_UNKNOWN,
            ocr_text="",
        )
        known_invoice = self.create_invoice(
            title="REGULAR INVOICE DETAIL OK",
            document_type=Invoice.DOCUMENT_TYPE_INVOICE,
            ocr_text="OCR TEXT",
        )

        self.client.force_login(
            self.user
        )

        response = self.client.get(
            reverse(
                "invoice_bot_report_detail",
                kwargs={
                    "category": "unknown-document-type",
                },
            )
        )

        self.assertEqual(
            response.status_code,
            200,
        )
        self.assertContains(
            response,
            "Неизвестный тип документа",
        )
        self.assertContains(
            response,
            matching_invoice.title,
        )
        self.assertNotContains(
            response,
            without_ocr_invoice.title,
        )
        self.assertNotContains(
            response,
            known_invoice.title,
        )

    def test_unknown_category_returns_404(self):
        self.client.force_login(
            self.user
        )

        response = self.client.get(
            reverse(
                "invoice_bot_report_detail",
                kwargs={
                    "category": "unknown",
                },
            )
        )

        self.assertEqual(
            response.status_code,
            404,
        )


    def test_export_excel_returns_xlsx_for_category(self):
        matching_invoice = self.create_invoice(
            title="EXCEL WITHOUT DATE DETAIL INVOICE",
            planned_payment_date=None,
        )
        other_invoice = self.create_invoice(
            title="EXCEL WITH DATE DETAIL INVOICE",
        )

        self.client.force_login(
            self.user
        )

        response = self.client.get(
            reverse(
                "export_invoice_bot_report_excel",
                kwargs={
                    "category": "without-planned-payment-date",
                },
            )
        )

        self.assertEqual(
            response.status_code,
            200,
        )
        self.assertEqual(
            response[
                "Content-Type"
            ],
            (
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )
        self.assertIn(
            "invoice_bot_report_without-planned-payment-date",
            response[
                "Content-Disposition"
            ],
        )

        workbook = load_workbook(
            BytesIO(
                response.content
            )
        )
        sheet = workbook.active

        self.assertEqual(
            sheet["A1"].value,
            "ID",
        )
        self.assertEqual(
            sheet["B1"].value,
            "Название",
        )
        self.assertEqual(
            sheet["I1"].value,
            "Причины блокировки",
        )

        values = [
            row[1]
            for row in sheet.iter_rows(
                min_row=2,
                values_only=True,
            )
        ]

        self.assertIn(
            matching_invoice.title,
            values,
        )
        self.assertNotIn(
            other_invoice.title,
            values,
        )

    def test_export_excel_unknown_category_returns_404(self):
        self.client.force_login(
            self.user
        )

        response = self.client.get(
            reverse(
                "export_invoice_bot_report_excel",
                kwargs={
                    "category": "unknown",
                },
            )
        )

        self.assertEqual(
            response.status_code,
            404,
        )


    def test_without_planned_payment_date_page_shows_quick_date_form(self):
        invoice = self.create_invoice(
            title="QUICK DATE FORM INVOICE",
            planned_payment_date=None,
        )

        self.client.force_login(
            self.user
        )

        response = self.client.get(
            reverse(
                "invoice_bot_report_detail",
                kwargs={
                    "category": "without-planned-payment-date",
                },
            )
        )

        self.assertEqual(
            response.status_code,
            200,
        )
        self.assertContains(
            response,
            reverse(
                "update_invoice_bot_report_planned_payment_date",
                kwargs={
                    "category": "without-planned-payment-date",
                    "invoice_id": invoice.id,
                },
            ),
        )
        self.assertContains(
            response,
            "Сохранить дату",
        )

    def test_quick_date_update_saves_date_and_removes_invoice_from_category(self):
        invoice = self.create_invoice(
            title="QUICK DATE UPDATE INVOICE",
            planned_payment_date=None,
        )

        self.client.force_login(
            self.user
        )

        response = self.client.post(
            reverse(
                "update_invoice_bot_report_planned_payment_date",
                kwargs={
                    "category": "without-planned-payment-date",
                    "invoice_id": invoice.id,
                },
            ),
            {
                "planned_payment_date": "2026-07-30",
            },
        )

        self.assertRedirects(
            response,
            reverse(
                "invoice_bot_report_detail",
                kwargs={
                    "category": "without-planned-payment-date",
                },
            ),
        )

        invoice.refresh_from_db()

        self.assertEqual(
            invoice.planned_payment_date,
            date(
                2026,
                7,
                30,
            ),
        )

        response = self.client.get(
            reverse(
                "invoice_bot_report_detail",
                kwargs={
                    "category": "without-planned-payment-date",
                },
            )
        )

        self.assertNotContains(
            response,
            invoice.title,
        )

    def test_quick_date_update_rejects_invalid_date(self):
        invoice = self.create_invoice(
            title="QUICK DATE INVALID INVOICE",
            planned_payment_date=None,
        )

        self.client.force_login(
            self.user
        )

        response = self.client.post(
            reverse(
                "update_invoice_bot_report_planned_payment_date",
                kwargs={
                    "category": "without-planned-payment-date",
                    "invoice_id": invoice.id,
                },
            ),
            {
                "planned_payment_date": "bad-date",
            },
        )

        self.assertRedirects(
            response,
            reverse(
                "invoice_bot_report_detail",
                kwargs={
                    "category": "without-planned-payment-date",
                },
            ),
        )

        invoice.refresh_from_db()

        self.assertIsNone(
            invoice.planned_payment_date,
        )

    def test_quick_date_update_is_not_available_for_other_categories(self):
        invoice = self.create_invoice(
            title="QUICK DATE WRONG CATEGORY INVOICE",
            planned_payment_date=None,
        )

        self.client.force_login(
            self.user
        )

        response = self.client.post(
            reverse(
                "update_invoice_bot_report_planned_payment_date",
                kwargs={
                    "category": "not-ready",
                    "invoice_id": invoice.id,
                },
            ),
            {
                "planned_payment_date": "2026-07-30",
            },
        )

        self.assertEqual(
            response.status_code,
            404,
        )


    def test_unverified_amount_page_shows_confirm_amount_button(self):
        invoice = self.create_invoice(
            title="QUICK AMOUNT FORM INVOICE",
            amount_verified=False,
        )

        self.client.force_login(
            self.user
        )

        response = self.client.get(
            reverse(
                "invoice_bot_report_detail",
                kwargs={
                    "category": "unverified-amount",
                },
            )
        )

        self.assertEqual(
            response.status_code,
            200,
        )
        self.assertContains(
            response,
            reverse(
                "confirm_invoice_bot_report_amount",
                kwargs={
                    "category": "unverified-amount",
                    "invoice_id": invoice.id,
                },
            ),
        )
        self.assertContains(
            response,
            "Подтвердить сумму",
        )

    def test_confirm_amount_sets_amount_verified_and_removes_invoice_from_category(self):
        invoice = self.create_invoice(
            title="QUICK AMOUNT UPDATE INVOICE",
            amount_verified=False,
            ocr_verified=False,
        )

        self.client.force_login(
            self.user
        )

        response = self.client.post(
            reverse(
                "confirm_invoice_bot_report_amount",
                kwargs={
                    "category": "unverified-amount",
                    "invoice_id": invoice.id,
                },
            )
        )

        self.assertRedirects(
            response,
            reverse(
                "invoice_bot_report_detail",
                kwargs={
                    "category": "unverified-amount",
                },
            ),
        )

        invoice.refresh_from_db()

        self.assertTrue(
            invoice.amount_verified,
        )
        self.assertFalse(
            invoice.ocr_verified,
        )
        self.assertIn(
            "Сумма подтверждена вручную из отчёта бота",
            invoice.ocr_comment,
        )

        response = self.client.get(
            reverse(
                "invoice_bot_report_detail",
                kwargs={
                    "category": "unverified-amount",
                },
            )
        )

        self.assertNotContains(
            response,
            invoice.title,
        )

    def test_confirm_amount_is_not_available_for_other_categories(self):
        invoice = self.create_invoice(
            title="QUICK AMOUNT WRONG CATEGORY INVOICE",
            amount_verified=False,
        )

        self.client.force_login(
            self.user
        )

        response = self.client.post(
            reverse(
                "confirm_invoice_bot_report_amount",
                kwargs={
                    "category": "not-ready",
                    "invoice_id": invoice.id,
                },
            )
        )

        self.assertEqual(
            response.status_code,
            404,
        )

        invoice.refresh_from_db()

        self.assertFalse(
            invoice.amount_verified,
        )


    def test_without_ocr_text_page_shows_retry_ocr_button(self):
        invoice = self.create_invoice(
            title="OCR RETRY FORM INVOICE",
            ocr_text="",
        )

        self.client.force_login(
            self.user
        )

        response = self.client.get(
            reverse(
                "invoice_bot_report_detail",
                kwargs={
                    "category": "without-ocr-text",
                },
            )
        )

        self.assertEqual(
            response.status_code,
            200,
        )
        self.assertContains(
            response,
            reverse(
                "retry_invoice_bot_report_ocr",
                kwargs={
                    "category": "without-ocr-text",
                    "invoice_id": invoice.id,
                },
            ),
        )
        self.assertContains(
            response,
            "Повторить OCR",
        )

    def test_retry_ocr_creates_pending_job_from_bot_report(self):
        invoice = self.create_invoice(
            title="OCR RETRY CREATE JOB INVOICE",
            ocr_text="",
        )

        self.client.force_login(
            self.user
        )

        response = self.client.post(
            reverse(
                "retry_invoice_bot_report_ocr",
                kwargs={
                    "category": "without-ocr-text",
                    "invoice_id": invoice.id,
                },
            )
        )

        self.assertRedirects(
            response,
            reverse(
                "invoice_bot_report_detail",
                kwargs={
                    "category": "without-ocr-text",
                },
            ),
        )

        job = OCRJob.objects.get(
            invoice=invoice
        )

        self.assertEqual(
            job.status,
            OCRJob.STATUS_PENDING,
        )
        self.assertEqual(
            job.source,
            OCRJob.SOURCE_SINGLE,
        )
        self.assertEqual(
            job.user,
            self.user,
        )

    def test_retry_ocr_does_not_create_duplicate_active_job(self):
        invoice = self.create_invoice(
            title="OCR RETRY DUPLICATE JOB INVOICE",
            ocr_text="",
        )

        OCRJob.objects.create(
            invoice=invoice,
            user=self.user,
            status=OCRJob.STATUS_PENDING,
            source=OCRJob.SOURCE_SINGLE,
        )

        self.client.force_login(
            self.user
        )

        response = self.client.post(
            reverse(
                "retry_invoice_bot_report_ocr",
                kwargs={
                    "category": "without-ocr-text",
                    "invoice_id": invoice.id,
                },
            )
        )

        self.assertRedirects(
            response,
            reverse(
                "invoice_bot_report_detail",
                kwargs={
                    "category": "without-ocr-text",
                },
            ),
        )

        self.assertEqual(
            OCRJob.objects.filter(
                invoice=invoice
            ).count(),
            1,
        )

    def test_retry_ocr_is_not_available_for_other_categories(self):
        invoice = self.create_invoice(
            title="OCR RETRY WRONG CATEGORY INVOICE",
            ocr_text="",
        )

        self.client.force_login(
            self.user
        )

        response = self.client.post(
            reverse(
                "retry_invoice_bot_report_ocr",
                kwargs={
                    "category": "not-ready",
                    "invoice_id": invoice.id,
                },
            )
        )

        self.assertEqual(
            response.status_code,
            404,
        )

        self.assertFalse(
            OCRJob.objects.filter(
                invoice=invoice
            ).exists()
        )
