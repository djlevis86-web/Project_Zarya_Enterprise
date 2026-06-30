from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO, StringIO
import csv

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import redirect
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models import (
    CompanyRequisites,
    Counterparty,
    Invoice,
    InvoicePayment,
    PaymentRegistry,
    PaymentRegistryItem,
)
from ..payment_registry_permissions import (
    require_payment_registry_permission,
    user_can_export_payment_registry,
)
from ..payment_registry_services import (
    check_payment_registry,
    recalculate_payment_registry,
)
from .payment_registry_helpers import (
    apply_payment_status_filter,
    apply_positive_payment_balance_filter,
)


@login_required
@require_payment_registry_permission(
    user_can_export_payment_registry,
    'Нет прав на выгрузку реестра оплаты.',
)
def export_payment_registry_draft_excel(request, registry_id):

    if request.method != 'POST':

        messages.warning(
            request,
            'Выгрузка реестра выполняется только из формы.'
        )

        return redirect(
            'payment_registry'
        )






    registry = (
        PaymentRegistry.objects
        .filter(
            id=registry_id,
            created_by=request.user,
            status=PaymentRegistry.STATUS_DRAFT,
        )
        .first()
    )

    if not registry:

        messages.warning(
            request,
            'Черновик реестра не найден или уже выгружен.'
        )

        return redirect(
            'payment_registry'
        )

    check_result = check_payment_registry(
        registry
    )

    if check_result['items_count'] == 0:

        messages.warning(
            request,
            f'Реестр №{registry.id} пуст. Сначала добавь счета.'
        )

        return redirect(
            'payment_registry'
        )

    if check_result['errors_count']:

        messages.warning(
            request,
            f'Реестр №{registry.id} нельзя выгрузить: ошибок {check_result["errors_count"]}.'
        )

        return redirect(
            'payment_registry'
        )

    items = (
        registry.items
        .select_related(
            'invoice',
            'invoice__counterparty',
            'invoice__user',
        )
        .exclude(
            status=PaymentRegistryItem.STATUS_CANCELLED
        )
        .order_by(
            'planned_payment_date',
            'invoice_id',
        )
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f'Registry {registry.id}'

    headers = [
        '№ реестра',
        'ID счета',
        'Номер счета',
        'Контрагент',
        'ИНН',
        'КПП',
        'Банк',
        'Расчетный счет',
        'БИК',
        'Корр. счет',
        'Сумма',
        'Дата оплаты',
        'Назначение платежа',
        'Ответственный',
    ]

    ws.append(headers)

    header_fill = PatternFill(
        fill_type='solid',
        fgColor='1F2937',
    )

    for cell in ws[1]:
        cell.font = Font(
            bold=True,
            color='FFFFFF',
        )
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal='center',
            vertical='center',
        )

    for item in items:

        invoice = item.invoice
        counterparty = invoice.counterparty

        amount = item.amount or Decimal('0')
        payment_date = item.planned_payment_date or invoice.planned_payment_date

        purpose = (
            getattr(invoice, 'payment_purpose', '')
            or getattr(invoice, 'purpose', '')
            or getattr(invoice, 'description', '')
            or ''
        )

        ws.append(
            [
                registry.id,
                invoice.id,
                invoice.invoice_number or '',
                counterparty.name if counterparty else '',
                getattr(counterparty, 'inn', '') if counterparty else '',
                getattr(counterparty, 'kpp', '') if counterparty else '',
                getattr(counterparty, 'bank_name', '') if counterparty else '',
                getattr(counterparty, 'account_number', '') if counterparty else '',
                getattr(counterparty, 'bik', '') if counterparty else '',
                getattr(counterparty, 'correspondent_account', '') if counterparty else '',
                float(amount),
                payment_date.strftime('%d.%m.%Y') if payment_date else '',
                purpose,
                invoice.user.get_username() if invoice.user else '',
            ]
        )

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:
            value = str(cell.value or '')
            max_length = max(
                max_length,
                len(value),
            )

        ws.column_dimensions[column_letter].width = min(
            max_length + 2,
            42,
        )

    ws.freeze_panes = 'A2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    now = timezone.now()

    registry.status = PaymentRegistry.STATUS_EXPORTED
    registry.exported_by = request.user
    registry.exported_at = now
    registry.save(
        update_fields=(
            'status',
            'exported_by',
            'exported_at',
        )
    )

    items.update(
        status=PaymentRegistryItem.STATUS_EXPORTED,
        exported_at=now,
    )

    recalculate_payment_registry(
        registry
    )

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    response['Content-Disposition'] = (
        f'attachment; filename="payment_registry_{registry.id}.xlsx"'
    )

    return response

@login_required
@require_payment_registry_permission(
    user_can_export_payment_registry,
    'Нет прав на выгрузку реестра оплаты.',
)
def export_payment_registry_draft_1c(request, registry_id):

    if request.method != 'POST':

        messages.warning(
            request,
            'Выгрузка реестра выполняется только из формы.'
        )

        return redirect(
            'payment_registry'
        )





    registry = (
        PaymentRegistry.objects
        .filter(
            id=registry_id,
            created_by=request.user,
            status=PaymentRegistry.STATUS_DRAFT,
        )
        .first()
    )

    if not registry:

        messages.warning(
            request,
            'Черновик реестра не найден или уже выгружен.'
        )

        return redirect(
            'payment_registry'
        )

    check_result = check_payment_registry(
        registry
    )

    if check_result['items_count'] == 0:

        messages.warning(
            request,
            f'Реестр №{registry.id} пуст. Сначала добавь счета.'
        )

        return redirect(
            'payment_registry'
        )

    if check_result['errors_count']:

        messages.warning(
            request,
            f'Реестр №{registry.id} нельзя выгрузить: ошибок {check_result["errors_count"]}.'
        )

        return redirect(
            'payment_registry'
        )

    items = (
        registry.items
        .select_related(
            'invoice',
            'invoice__counterparty',
            'invoice__user',
        )
        .exclude(
            status=PaymentRegistryItem.STATUS_CANCELLED
        )
        .order_by(
            'planned_payment_date',
            'invoice_id',
        )
    )

    buffer = StringIO()
    writer = csv.writer(
        buffer,
        delimiter=';',
        lineterminator='\n',
    )

    writer.writerow(
        [
            'RegistryID',
            'InvoiceID',
            'InvoiceNumber',
            'Counterparty',
            'INN',
            'KPP',
            'Bank',
            'BankAccount',
            'BIK',
            'CorrAccount',
            'Amount',
            'PaymentDate',
            'Purpose',
        ]
    )

    for item in items:

        invoice = item.invoice
        counterparty = invoice.counterparty
        payment_date = item.planned_payment_date or invoice.planned_payment_date

        purpose = (
            getattr(invoice, 'payment_purpose', '')
            or getattr(invoice, 'purpose', '')
            or getattr(invoice, 'description', '')
            or ''
        )

        writer.writerow(
            [
                registry.id,
                invoice.id,
                invoice.invoice_number or '',
                counterparty.name if counterparty else '',
                getattr(counterparty, 'inn', '') if counterparty else '',
                getattr(counterparty, 'kpp', '') if counterparty else '',
                getattr(counterparty, 'bank_name', '') if counterparty else '',
                getattr(counterparty, 'account_number', '') if counterparty else '',
                getattr(counterparty, 'bik', '') if counterparty else '',
                getattr(counterparty, 'correspondent_account', '') if counterparty else '',
                str(item.amount).replace('.', ','),
                payment_date.strftime('%d.%m.%Y') if payment_date else '',
                purpose,
            ]
        )

    now = timezone.now()

    registry.status = PaymentRegistry.STATUS_EXPORTED
    registry.exported_by = request.user
    registry.exported_at = now
    registry.save(
        update_fields=(
            'status',
            'exported_by',
            'exported_at',
        )
    )

    items.update(
        status=PaymentRegistryItem.STATUS_EXPORTED,
        exported_at=now,
    )

    recalculate_payment_registry(
        registry
    )

    content = '\ufeff' + buffer.getvalue()

    response = HttpResponse(
        content,
        content_type='text/plain; charset=utf-8',
    )

    response['Content-Disposition'] = (
        f'attachment; filename="payment_registry_{registry.id}_1c.txt"'
    )

    return response

@login_required
@require_payment_registry_permission(
    user_can_export_payment_registry,
    'Нет прав на выгрузку реестра оплаты.',
)
def export_payment_registry_excel(request):

    selected_status = request.GET.get(
        'status',
        Invoice.STATUS_APPROVED
    )

    selected_counterparty = request.GET.get(
        'counterparty',
        ''
    )

    search_query = request.GET.get(
        'q',
        ''
    ).strip()

    date_from = request.GET.get(
        'date_from',
        ''
    ).strip()

    date_to = request.GET.get(
        'date_to',
        ''
    ).strip()

    invoices = (
        Invoice.objects
        .select_related(
            'counterparty',
            'user'
        )
        .exclude(
            status=Invoice.STATUS_PAID
        )
    )

    if selected_status and selected_status != 'all':

        invoices = invoices.filter(
            status=selected_status
        )

    if selected_counterparty:

        invoices = invoices.filter(
            counterparty_id=selected_counterparty
        )

    if search_query:

        invoices = invoices.filter(
            Q(title__icontains=search_query)
            |
            Q(original_filename__icontains=search_query)
            |
            Q(invoice_number__icontains=search_query)
            |
            Q(vendor__icontains=search_query)
            |
            Q(counterparty__name__icontains=search_query)
        )

    if date_from:

        invoices = invoices.filter(
            planned_payment_date__gte=date_from
        )

    if date_to:

        invoices = invoices.filter(
            planned_payment_date__lte=date_to
        )

    invoices = invoices.order_by(
        'planned_payment_date',
        '-payment_priority',
        'counterparty__name',
        'id'
    )

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = 'Реестр оплаты'

    headers = [
        'ID',
        'Контрагент',
        'ИНН',
        'КПП',
        'Номер счета',
        'Дата счета',
        'Сумма',
        'Плановая дата оплаты',
        'Приоритет',
        'Статус',
        'Назначение платежа',
    ]

    sheet.append(
        headers
    )

    header_fill = PatternFill(
        fill_type='solid',
        fgColor='E5E7EB'
    )

    for cell in sheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.fill = header_fill

        cell.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )

    total_amount = Decimal(
        '0.00'
    )

    for invoice in invoices:

        if invoice.counterparty:

            counterparty_name = invoice.counterparty.name or ''
            inn = invoice.counterparty.inn or ''
            kpp = invoice.counterparty.kpp or ''

        else:

            counterparty_name = invoice.vendor or ''
            inn = ''
            kpp = ''

        amount = invoice.amount or Decimal(
            '0.00'
        )

        total_amount += amount

        if invoice.invoice_number:

            payment_purpose = (
                f'Оплата по счету №{invoice.invoice_number}'
            )

            if invoice.invoice_date:

                payment_purpose += (
                    f' от {invoice.invoice_date}'
                )

        else:

            payment_purpose = (
                f'Оплата по счету ID {invoice.id}'
            )

        sheet.append(
            [
                invoice.id,
                counterparty_name,
                inn,
                kpp,
                invoice.invoice_number or '',
                invoice.invoice_date or '',
                amount,
                (
                    invoice.planned_payment_date.strftime(
                        '%d.%m.%Y'
                    )
                    if invoice.planned_payment_date
                    else ''
                ),
                invoice.payment_priority,
                invoice.get_status_display(),
                payment_purpose,
            ]
        )

    total_row = sheet.max_row + 2

    sheet.cell(
        row=total_row,
        column=6,
        value='Итого:'
    )

    sheet.cell(
        row=total_row,
        column=7,
        value=total_amount
    )

    sheet.cell(
        row=total_row,
        column=6
    ).font = Font(
        bold=True
    )

    sheet.cell(
        row=total_row,
        column=7
    ).font = Font(
        bold=True
    )

    for row in sheet.iter_rows():

        for cell in row:

            cell.alignment = Alignment(
                vertical='top',
                wrap_text=True
            )

    column_widths = {
        'A': 8,
        'B': 36,
        'C': 16,
        'D': 14,
        'E': 18,
        'F': 16,
        'G': 16,
        'H': 20,
        'I': 12,
        'J': 18,
        'K': 48,
    }

    for column_letter, width in column_widths.items():

        sheet.column_dimensions[
            column_letter
        ].width = width

    for row_number in range(
        2,
        sheet.max_row + 1
    ):

        sheet.cell(
            row=row_number,
            column=7
        ).number_format = '#,##0.00'

    sheet.freeze_panes = 'A2'

    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet'
        )
    )

    response[
        'Content-Disposition'
    ] = (
        'attachment; filename="payment_registry.xlsx"'
    )

    workbook.save(
        response
    )

    return response

@login_required
@require_payment_registry_permission(
    user_can_export_payment_registry,
    'Нет прав на выгрузку реестра оплаты.',
)
def export_payment_registry_1c(request):

    company = CompanyRequisites.objects.first()

    if not company:

        messages.error(
            request,
            'Сначала заполните реквизиты организации в админке.'
        )

        return redirect(
            'payment_registry'
        )

    selected_status = request.GET.get(
        'status',
        Invoice.STATUS_APPROVED
    )

    selected_counterparty = request.GET.get(
        'counterparty',
        ''
    )

    search_query = request.GET.get(
        'q',
        ''
    ).strip()

    date_from = request.GET.get(
        'date_from',
        ''
    ).strip()

    date_to = request.GET.get(
        'date_to',
        ''
    ).strip()

    invoices = (
        Invoice.objects
        .select_related(
            'counterparty',
            'user'
        )
        .exclude(
            status=Invoice.STATUS_PAID
        )
    )

    if selected_status and selected_status != 'all':

        invoices = invoices.filter(
            status=selected_status
        )

    if selected_counterparty:

        invoices = invoices.filter(
            counterparty_id=selected_counterparty
        )

    if search_query:

        invoices = invoices.filter(
            Q(title__icontains=search_query)
            |
            Q(original_filename__icontains=search_query)
            |
            Q(invoice_number__icontains=search_query)
            |
            Q(vendor__icontains=search_query)
            |
            Q(counterparty__name__icontains=search_query)
        )

    if date_from:

        invoices = invoices.filter(
            planned_payment_date__gte=date_from
        )

    if date_to:

        invoices = invoices.filter(
            planned_payment_date__lte=date_to
        )

    invoices = invoices.order_by(
        'planned_payment_date',
        '-payment_priority',
        'counterparty__name',
        'id'
    )

    missing_requisites = []

    for invoice in invoices:

        counterparty = invoice.counterparty

        if not counterparty:

            missing_requisites.append(
                f'Счет #{invoice.id}: контрагент не найден'
            )

            continue

        required_values = [
            counterparty.inn,
            counterparty.bank_name,
            counterparty.bik,
            counterparty.account_number,
        ]

        if any(
            not value
            for value in required_values
        ):

            missing_requisites.append(
                f'Счет #{invoice.id}: {counterparty.name}'
            )

    if missing_requisites:

        messages.error(
            request,
            (
                'Выгрузка 1С TXT остановлена. '
                'Есть контрагенты без обязательных платежных реквизитов.'
            )
        )

        return redirect(
            'counterparties_missing_requisites'
        )

    def clean_value(value):

        if value is None:

            return ''

        value = str(value)

        value = value.replace(
            '\r',
            ' '
        )

        value = value.replace(
            '\n',
            ' '
        )

        value = value.strip()

        return value

    def format_date(value):

        if value:

            return value.strftime(
                '%d.%m.%Y'
            )

        return date.today().strftime(
            '%d.%m.%Y'
        )

    def format_amount(value):

        if not value:

            return '0.00'

        return f'{value:.2f}'

    created_at = datetime.now()

    lines = []

    lines.append(
        '1CClientBankExchange'
    )

    lines.append(
        'ВерсияФормата=1.03'
    )

    lines.append(
        'Кодировка=Windows'
    )

    lines.append(
        'Отправитель=Project Zarya'
    )

    lines.append(
        'Получатель=1С'
    )

    lines.append(
        f'ДатаСоздания={created_at.strftime("%d.%m.%Y")}'
    )

    lines.append(
        f'ВремяСоздания={created_at.strftime("%H:%M:%S")}'
    )

    lines.append(
        f'РасчСчет={clean_value(company.account_number)}'
    )

    for invoice in invoices:

        if not invoice.counterparty:

            continue

        counterparty = invoice.counterparty

        amount = invoice.amount or 0

        payment_date = invoice.planned_payment_date or date.today()

        if invoice.invoice_number:

            payment_purpose = (
                f'Оплата по счету №{invoice.invoice_number}'
            )

            if invoice.invoice_date:

                payment_purpose += (
                    f' от {invoice.invoice_date}'
                )

        else:

            payment_purpose = (
                f'Оплата по счету ID {invoice.id}'
            )

        lines.append(
            'СекцияДокумент=Платежное поручение'
        )

        lines.append(
            f'Номер={invoice.id}'
        )

        lines.append(
            f'Дата={format_date(payment_date)}'
        )

        lines.append(
            f'Сумма={format_amount(amount)}'
        )

        lines.append(
            f'Плательщик={clean_value(company.name)}'
        )

        lines.append(
            f'ПлательщикИНН={clean_value(company.inn)}'
        )

        lines.append(
            f'ПлательщикКПП={clean_value(company.kpp)}'
        )

        lines.append(
            f'ПлательщикСчет={clean_value(company.account_number)}'
        )

        lines.append(
            f'ПлательщикБанк1={clean_value(company.bank_name)}'
        )

        lines.append(
            f'ПлательщикБИК={clean_value(company.bik)}'
        )

        lines.append(
            f'ПлательщикКорсчет={clean_value(company.correspondent_account)}'
        )

        lines.append(
            f'Получатель={clean_value(counterparty.name)}'
        )

        lines.append(
            f'ПолучательИНН={clean_value(counterparty.inn)}'
        )

        lines.append(
            f'ПолучательКПП={clean_value(counterparty.kpp)}'
        )

        lines.append(
            f'ПолучательСчет={clean_value(counterparty.account_number)}'
        )

        lines.append(
            f'ПолучательБанк1={clean_value(counterparty.bank_name)}'
        )

        lines.append(
            f'ПолучательБИК={clean_value(counterparty.bik)}'
        )

        lines.append(
            f'ПолучательКорсчет={clean_value(counterparty.correspondent_account)}'
        )

        lines.append(
            'ВидПлатежа=Электронно'
        )

        lines.append(
            'ВидОплаты=01'
        )

        lines.append(
            'Очередность=5'
        )

        lines.append(
            f'НазначениеПлатежа={clean_value(payment_purpose)}'
        )

        lines.append(
            'КонецДокумента'
        )

    lines.append(
        'КонецФайла'
    )

    content = '\r\n'.join(
        lines
    )

    response = HttpResponse(
        content.encode(
            'cp1251',
            errors='replace'
        ),
        content_type='text/plain; charset=windows-1251'
    )

    response[
        'Content-Disposition'
    ] = (
        'attachment; filename="payment_registry_1c.txt"'
    )

    return response
