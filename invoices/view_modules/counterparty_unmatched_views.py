from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from ..counterparty_service import extract_requisites_near_vendor, normalize_counterparty_name
from ..models import Invoice


@login_required
def unmatched_counterparties(request):

    invoices = (
        Invoice.objects
        .select_related(
            'user'
        )
        .filter(
            counterparty_match_status=Invoice.COUNTERPARTY_MATCH_NOT_FOUND
        )
        .order_by(
            'vendor',
            'id'
        )
    )

    groups_map = {}

    total_amount = Decimal(
        '0.00'
    )

    invoices_count = 0

    for invoice in invoices:

        vendor_name = invoice.vendor or 'Не определен'

        group_key = vendor_name.strip().upper()

        if group_key not in groups_map:

            groups_map[group_key] = {
                'vendor_name': vendor_name,
                'invoices': [],
                'invoice_numbers': [],
                'count': 0,
                'total_amount': Decimal('0.00'),
            }

        amount = invoice.amount or Decimal(
            '0.00'
        )

        groups_map[group_key]['invoices'].append(
            invoice
        )

        groups_map[group_key]['count'] += 1

        groups_map[group_key]['total_amount'] += amount

        if invoice.invoice_number:

            groups_map[group_key]['invoice_numbers'].append(
                invoice.invoice_number
            )

        total_amount += amount

        invoices_count += 1

    groups = sorted(
        groups_map.values(),
        key=lambda item: item['vendor_name']
    )

    return render(
        request,
        'invoices/unmatched_counterparties.html',
        {
            'groups': groups,
            'invoices_count': invoices_count,
            'groups_count': len(groups),
            'total_amount': total_amount,
        }
    )

@login_required
def export_unmatched_counterparties_excel(request):

    invoices = (
        Invoice.objects
        .filter(
            counterparty_match_status=Invoice.COUNTERPARTY_MATCH_NOT_FOUND
        )
        .exclude(
            vendor__isnull=True
        )
        .exclude(
            vendor=''
        )
        .order_by(
            'vendor',
            'id'
        )
    )

    candidates = {}

    for invoice in invoices:

        vendor_name = normalize_counterparty_name(
            invoice.vendor
        )

        if not vendor_name:

            continue

        inn, kpp = extract_requisites_near_vendor(
            invoice.ocr_text or '',
            vendor_name
        )

        key = (
            inn or '',
            kpp or '',
            vendor_name.upper()
        )

        if key not in candidates:

            candidates[key] = {
                'name': vendor_name,
                'inn': inn or '',
                'kpp': kpp or '',
                'invoice_ids': [],
                'invoice_numbers': [],
                'amount_total': Decimal('0.00'),
                'count': 0,
            }

        candidates[key]['invoice_ids'].append(
            str(invoice.id)
        )

        if invoice.invoice_number:

            candidates[key]['invoice_numbers'].append(
                str(invoice.invoice_number)
            )

        candidates[key]['amount_total'] += (
            invoice.amount or Decimal('0.00')
        )

        candidates[key]['count'] += 1

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = 'Кандидаты 1С'

    headers = [
        'Код',
        'Наименование',
        'Полное наименование',
        'ИНН',
        'КПП',
        'Банк',
        'БИК',
        'Расчетный счет',
        'Корреспондентский счет',
        'Активен',
        'Пометка удаления',
        'Количество счетов',
        'ID счетов',
        'Номера счетов',
        'Сумма по счетам',
        'Комментарий',
    ]

    sheet.append(
        headers
    )

    header_fill = PatternFill(
        fill_type='solid',
        fgColor='E5E7EB'
    )

    for cell in sheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.fill = header_fill

        cell.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )

    for candidate in sorted(
        candidates.values(),
        key=lambda item: item['name']
    ):

        sheet.append(
            [
                '',
                candidate['name'],
                candidate['name'],
                candidate['inn'],
                candidate['kpp'],
                '',
                '',
                '',
                '',
                'Да',
                '',
                candidate['count'],
                ', '.join(
                    candidate['invoice_ids']
                ),
                ', '.join(
                    sorted(
                        set(
                            candidate['invoice_numbers']
                        )
                    )
                ),
                candidate['amount_total'],
                'Сформировано из OCR. Нужно сверить с 1С.',
            ]
        )

    for row in sheet.iter_rows():

        for cell in row:

            cell.alignment = Alignment(
                vertical='top',
                wrap_text=True
            )

    column_widths = {
        'A': 18,
        'B': 42,
        'C': 48,
        'D': 16,
        'E': 14,
        'F': 34,
        'G': 14,
        'H': 24,
        'I': 24,
        'J': 12,
        'K': 18,
        'L': 16,
        'M': 28,
        'N': 28,
        'O': 18,
        'P': 42,
    }

    for column_letter, width in column_widths.items():

        sheet.column_dimensions[
            column_letter
        ].width = width

    for row_number in range(
        2,
        sheet.max_row + 1
    ):

        sheet.cell(
            row=row_number,
            column=15
        ).number_format = '#,##0.00'

    sheet.freeze_panes = 'A2'

    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet'
        )
    )

    response[
        'Content-Disposition'
    ] = (
        'attachment; filename="counterparty_candidates_1c.xlsx"'
    )

    workbook.save(
        response
    )

    return response
