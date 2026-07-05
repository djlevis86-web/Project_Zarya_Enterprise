from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_POST

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..bot_report_services import (
    BOT_REPORT_CATEGORY_UNVERIFIED_AMOUNT,
    BOT_REPORT_CATEGORY_WITHOUT_OCR_TEXT,
    BOT_REPORT_CATEGORY_WITHOUT_PLANNED_PAYMENT_DATE,
    get_invoice_bot_report_category,
    get_invoice_bot_report_items,
)
from ..log_service import create_invoice_log
from ..models import Invoice, OCRJob
from ..payment_registry_permissions import (
    require_payment_registry_permission,
    user_can_manage_payment_registry,
)


EXCEL_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument."
    "spreadsheetml.sheet"
)


@login_required
@require_payment_registry_permission(
    user_can_manage_payment_registry,
    "Нет прав на просмотр отчёта бота.",
)
def invoice_bot_report_detail(request, category):
    category_data, invoice_items = get_invoice_bot_report_items(
        category
    )

    if category_data is None:
        raise Http404(
            "Категория отчёта бота не найдена."
        )

    _attach_ocr_job_diagnostics(
        category,
        invoice_items,
    )

    paginator = Paginator(
        invoice_items,
        50,
    )

    page_obj = paginator.get_page(
        request.GET.get(
            "page"
        )
    )

    return render(
        request,
        "invoices/invoice_bot_report_detail.html",
        {
            "category": category,
            "category_data": category_data,
            "page_obj": page_obj,
            "items_count": len(
                invoice_items
            ),
        }
    )


def _attach_ocr_job_diagnostics(category, invoice_items):
    if category != BOT_REPORT_CATEGORY_WITHOUT_OCR_TEXT:
        return

    invoice_ids = [
        item["invoice"].id
        for item in invoice_items
    ]

    if not invoice_ids:
        return

    latest_jobs_by_invoice_id = {}

    latest_jobs = (
        OCRJob.objects
        .filter(
            invoice_id__in=invoice_ids,
        )
        .select_related(
            "user",
        )
        .order_by(
            "invoice_id",
            "-created_at",
            "-id",
        )
    )

    for job in latest_jobs:
        if job.invoice_id not in latest_jobs_by_invoice_id:
            latest_jobs_by_invoice_id[job.invoice_id] = job

    active_invoice_ids = set(
        OCRJob.objects
        .filter(
            invoice_id__in=invoice_ids,
            status__in=[
                OCRJob.STATUS_PENDING,
                OCRJob.STATUS_PROCESSING,
            ],
        )
        .values_list(
            "invoice_id",
            flat=True,
        )
    )

    for item in invoice_items:
        invoice = item["invoice"]
        item["latest_ocr_job"] = latest_jobs_by_invoice_id.get(
            invoice.id
        )
        item["has_active_ocr_job"] = invoice.id in active_invoice_ids


@login_required
@require_POST
@require_payment_registry_permission(
    user_can_manage_payment_registry,
    "Нет прав на быстрое исправление отчёта бота.",
)
def update_invoice_bot_report_planned_payment_date(
    request,
    category,
    invoice_id,
):
    category_data = get_invoice_bot_report_category(
        category
    )

    if category_data is None:
        raise Http404(
            "Категория отчёта бота не найдена."
        )

    if category != BOT_REPORT_CATEGORY_WITHOUT_PLANNED_PAYMENT_DATE:
        raise Http404(
            "Быстрое исправление даты доступно только для категории без даты оплаты."
        )

    invoice = get_object_or_404(
        Invoice,
        id=invoice_id,
        is_deleted=False,
    )

    planned_payment_date_raw = request.POST.get(
        "planned_payment_date",
        "",
    )

    if not planned_payment_date_raw:
        messages.error(
            request,
            "Укажите плановую дату оплаты."
        )

        return redirect(
            "invoice_bot_report_detail",
            category=category,
        )

    new_planned_payment_date = parse_date(
        planned_payment_date_raw
    )

    if not new_planned_payment_date:
        messages.error(
            request,
            "Введите корректную плановую дату оплаты."
        )

        return redirect(
            "invoice_bot_report_detail",
            category=category,
        )

    old_planned_payment_date = invoice.planned_payment_date

    if old_planned_payment_date == new_planned_payment_date:
        messages.info(
            request,
            f"По счёту #{invoice.id} дата оплаты не изменилась."
        )

        return redirect(
            "invoice_bot_report_detail",
            category=category,
        )

    invoice.planned_payment_date = new_planned_payment_date

    update_fields = [
        "planned_payment_date",
    ]

    if any(field.name == "updated_at" for field in invoice._meta.fields):
        update_fields.append(
            "updated_at"
        )

    invoice.save(
        update_fields=update_fields
    )

    old_date_display = (
        old_planned_payment_date.strftime("%d.%m.%Y")
        if old_planned_payment_date
        else "не указана"
    )
    new_date_display = new_planned_payment_date.strftime(
        "%d.%m.%Y"
    )

    create_invoice_log(
        invoice,
        request.user,
        (
            "Плановая дата оплаты изменена из отчёта бота: "
            f"{old_date_display} → {new_date_display}."
        )
    )

    messages.success(
        request,
        f"Плановая дата оплаты по счёту #{invoice.id} сохранена."
    )

    return redirect(
        "invoice_bot_report_detail",
        category=category,
    )


@login_required
@require_POST
@require_payment_registry_permission(
    user_can_manage_payment_registry,
    "Нет прав на быстрое подтверждение суммы из отчёта бота.",
)
def confirm_invoice_bot_report_amount(
    request,
    category,
    invoice_id,
):
    category_data = get_invoice_bot_report_category(
        category
    )

    if category_data is None:
        raise Http404(
            "Категория отчёта бота не найдена."
        )

    if category != BOT_REPORT_CATEGORY_UNVERIFIED_AMOUNT:
        raise Http404(
            "Быстрое подтверждение суммы доступно только для категории с неподтверждённой суммой."
        )

    invoice = get_object_or_404(
        Invoice,
        id=invoice_id,
        is_deleted=False,
    )

    if invoice.amount_verified:
        messages.info(
            request,
            f"Сумма по счёту #{invoice.id} уже подтверждена."
        )

        return redirect(
            "invoice_bot_report_detail",
            category=category,
        )

    invoice.amount_verified = True
    invoice.ocr_comment = (
        "Сумма подтверждена вручную из отчёта бота. "
        "OCR-статус не изменялся."
    )

    update_fields = [
        "amount_verified",
        "ocr_comment",
    ]

    if any(field.name == "updated_at" for field in invoice._meta.fields):
        update_fields.append(
            "updated_at"
        )

    invoice.save(
        update_fields=update_fields
    )

    create_invoice_log(
        invoice,
        request.user,
        "Сумма счёта подтверждена вручную из отчёта бота."
    )

    messages.success(
        request,
        f"Сумма по счёту #{invoice.id} подтверждена."
    )

    return redirect(
        "invoice_bot_report_detail",
        category=category,
    )


@login_required
@require_POST
@require_payment_registry_permission(
    user_can_manage_payment_registry,
    "Нет прав на повторный запуск OCR из отчёта бота.",
)
def retry_invoice_bot_report_ocr(
    request,
    category,
    invoice_id,
):
    category_data = get_invoice_bot_report_category(
        category
    )

    if category_data is None:
        raise Http404(
            "Категория отчёта бота не найдена."
        )

    if category != BOT_REPORT_CATEGORY_WITHOUT_OCR_TEXT:
        raise Http404(
            "Повтор OCR доступен только для категории без OCR-текста."
        )

    invoice = get_object_or_404(
        Invoice,
        id=invoice_id,
        is_deleted=False,
    )

    if invoice.ocr_text:
        messages.info(
            request,
            f"По счёту #{invoice.id} OCR-текст уже есть."
        )

        return redirect(
            "invoice_bot_report_detail",
            category=category,
        )

    existing_job = (
        OCRJob.objects
        .filter(
            invoice=invoice,
            status__in=[
                OCRJob.STATUS_PENDING,
                OCRJob.STATUS_PROCESSING,
            ],
        )
        .first()
    )

    if existing_job:
        messages.info(
            request,
            f"OCR по счёту #{invoice.id} уже стоит в очереди."
        )

        return redirect(
            "invoice_bot_report_detail",
            category=category,
        )

    ocr_job = OCRJob.objects.create(
        invoice=invoice,
        user=request.user,
        status=OCRJob.STATUS_PENDING,
        source=OCRJob.SOURCE_SINGLE,
        message="OCR поставлен в очередь из отчёта бота.",
    )

    create_invoice_log(
        invoice,
        request.user,
        f"OCR повторно поставлен в очередь из отчёта бота. OCRJob #{ocr_job.id}."
    )

    messages.success(
        request,
        f"OCR по счёту #{invoice.id} поставлен в очередь."
    )

    return redirect(
        "invoice_bot_report_detail",
        category=category,
    )


@login_required
@require_payment_registry_permission(
    user_can_manage_payment_registry,
    "Нет прав на экспорт отчёта бота.",
)
def export_invoice_bot_report_excel(request, category):
    category_data, invoice_items = get_invoice_bot_report_items(
        category
    )

    if category_data is None:
        raise Http404(
            "Категория отчёта бота не найдена."
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчёт бота"

    headers = [
        "ID",
        "Название",
        "Оригинальный файл",
        "Контрагент",
        "Сумма",
        "Статус",
        "Плановая дата оплаты",
        "Ответственный",
        "Причины блокировки",
        "Предупреждения",
    ]

    sheet.append(
        headers
    )

    for item in invoice_items:
        invoice = item["invoice"]

        sheet.append(
            [
                invoice.id,
                invoice.title,
                invoice.original_filename or "",
                _get_counterparty_name(
                    invoice
                ),
                invoice.amount,
                invoice.get_status_display(),
                _format_date(
                    invoice.planned_payment_date
                ),
                _get_invoice_user_name(
                    invoice
                ),
                "\n".join(
                    item["errors"]
                ),
                "\n".join(
                    item["warnings"]
                ),
            ]
        )

    _style_invoice_bot_report_sheet(
        sheet
    )

    response = HttpResponse(
        content_type=EXCEL_CONTENT_TYPE,
    )

    filename = (
        f"invoice_bot_report_{category}_"
        f"{timezone.localdate().isoformat()}.xlsx"
    )

    response[
        "Content-Disposition"
    ] = (
        f'attachment; filename="{filename}"'
    )

    workbook.save(
        response
    )

    return response


def _get_counterparty_name(invoice):
    if invoice.counterparty:
        return invoice.counterparty.name

    return ""


def _get_invoice_user_name(invoice):
    if not invoice.user_id:
        return ""

    full_name = invoice.user.get_full_name()

    if full_name:
        return full_name

    return invoice.user.username


def _format_date(value):
    if not value:
        return ""

    return value.strftime(
        "%d.%m.%Y"
    )


def _style_invoice_bot_report_sheet(sheet):
    header_fill = PatternFill(
        "solid",
        fgColor="1F2937",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )
    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    body_alignment = Alignment(
        vertical="top",
        wrap_text=True,
    )

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row in sheet.iter_rows(
        min_row=2,
    ):
        for cell in row:
            cell.alignment = body_alignment

    widths = [
        10,
        34,
        34,
        34,
        16,
        18,
        22,
        24,
        48,
        48,
    ]

    for index, width in enumerate(
        widths,
        start=1,
    ):
        sheet.column_dimensions[
            get_column_letter(
                index
            )
        ].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
