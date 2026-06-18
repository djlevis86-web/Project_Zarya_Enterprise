from pathlib import Path

from django.core.management.base import BaseCommand

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from invoices.models import Invoice
from invoices.counterparty_service import (
    normalize_counterparty_name,
    extract_requisites_near_vendor,
)


class Command(BaseCommand):

    help = 'Выгружает кандидатов контрагентов из OCR для сверки с 1С'

    def add_arguments(self, parser):

        parser.add_argument(
            'output_path',
            type=str,
            nargs='?',
            default='counterparty_candidates_1c.xlsx',
            help='Путь для сохранения XLSX файла'
        )

    def handle(self, *args, **options):

        output_path = Path(
            options['output_path']
        )

        invoices = (
            Invoice.objects
            .exclude(vendor__isnull=True)
            .exclude(vendor='')
            .order_by('vendor', 'id')
        )

        candidates = {}

        for invoice in invoices:

            vendor_name = normalize_counterparty_name(
                invoice.vendor
            )

            if not vendor_name:

                continue

            inn, kpp = extract_requisites_near_vendor(
                invoice.ocr_text or '',
                vendor_name
            )

            key = (
                inn or '',
                kpp or '',
                vendor_name.upper()
            )

            if key not in candidates:

                candidates[key] = {
                    'name': vendor_name,
                    'inn': inn or '',
                    'kpp': kpp or '',
                    'invoice_ids': [],
                    'invoice_numbers': [],
                    'amount_total': 0,
                    'count': 0,
                }

            candidates[key]['invoice_ids'].append(
                str(invoice.id)
            )

            if invoice.invoice_number:

                candidates[key]['invoice_numbers'].append(
                    str(invoice.invoice_number)
                )

            candidates[key]['amount_total'] += (
                invoice.amount or 0
            )

            candidates[key]['count'] += 1

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = 'Кандидаты 1С'

        headers = [
            'Код',
            'Наименование',
            'Полное наименование',
            'ИНН',
            'КПП',
            'Банк',
            'БИК',
            'Расчетный счет',
            'Корреспондентский счет',
            'Активен',
            'Пометка удаления',
            'Количество счетов',
            'ID счетов',
            'Номера счетов',
            'Сумма по счетам',
            'Комментарий',
        ]

        sheet.append(
            headers
        )

        header_fill = PatternFill(
            fill_type='solid',
            fgColor='E5E7EB'
        )

        for cell in sheet[1]:

            cell.font = Font(
                bold=True
            )

            cell.fill = header_fill

            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )

        for candidate in sorted(
            candidates.values(),
            key=lambda item: item['name']
        ):

            sheet.append(
                [
                    '',
                    candidate['name'],
                    candidate['name'],
                    candidate['inn'],
                    candidate['kpp'],
                    '',
                    '',
                    '',
                    '',
                    'Да',
                    '',
                    candidate['count'],
                    ', '.join(candidate['invoice_ids']),
                    ', '.join(sorted(set(candidate['invoice_numbers']))),
                    candidate['amount_total'],
                    'Сформировано из OCR, нужно сверить с 1С',
                ]
            )

        for row in sheet.iter_rows():

            for cell in row:

                cell.alignment = Alignment(
                    vertical='top',
                    wrap_text=True
                )

        column_widths = {
            'A': 18,
            'B': 42,
            'C': 48,
            'D': 16,
            'E': 14,
            'F': 34,
            'G': 14,
            'H': 24,
            'I': 24,
            'J': 12,
            'K': 18,
            'L': 16,
            'M': 28,
            'N': 28,
            'O': 18,
            'P': 42,
        }

        for column_letter, width in column_widths.items():

            sheet.column_dimensions[
                column_letter
            ].width = width

        for row_number in range(
            2,
            sheet.max_row + 1
        ):

            sheet.cell(
                row=row_number,
                column=15
            ).number_format = '#,##0.00'

        sheet.freeze_panes = 'A2'

        workbook.save(
            output_path
        )

        self.stdout.write(
            self.style.SUCCESS(
                f'Файл создан: {output_path}'
            )
        )

        self.stdout.write(
            f'Кандидатов: {len(candidates)}'
        )