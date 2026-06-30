from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO, StringIO
import csv

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import DecimalField, F, Q, Sum, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models import (
    CompanyRequisites,
    Counterparty,
    Invoice,
    InvoicePayment,
    PaymentRegistry,
    PaymentRegistryItem,
)
from ..payment_registry_permissions import (
    require_payment_registry_permission,
    user_can_cancel_payment_registry,
    user_can_check_payment_registry,
    user_can_export_payment_registry,
    user_can_manage_payment_registry,
    user_can_mark_payment_registry_paid,
)


def get_payment_registry_permission_context(user):
    return {
        "can_manage_payment_registry": user_can_manage_payment_registry(user),
        "can_check_payment_registry": user_can_check_payment_registry(user),
        "can_export_payment_registry": user_can_export_payment_registry(user),
        "can_mark_payment_registry_paid": user_can_mark_payment_registry_paid(user),
        "can_cancel_payment_registry": user_can_cancel_payment_registry(user),
    }


PAYMENT_STATUS_FILTER_CHOICES = (
    ("", "Все оплаты"),
    ("unpaid", "Не оплачен"),
    ("partial", "Частично оплачен"),
    ("paid", "Оплачен"),
    ("overpaid", "Переплата"),
)


OCR_STATUS_FILTER_CHOICES = (
    ("", "Все OCR-статусы"),
    ("verified", "Сумма подтверждена"),
    ("unverified", "Сумма требует проверки"),
)


def apply_payment_status_filter(queryset, payment_status):
    if not payment_status:
        return queryset

    queryset = queryset.annotate(
        payment_paid_sum=Sum(
            "payments__amount",
            filter=Q(
                payments__status=InvoicePayment.STATUS_POSTED
            )
        )
    )

    if payment_status == "unpaid":
        return queryset.filter(
            Q(payment_paid_sum__isnull=True)
            |
            Q(payment_paid_sum__lte=0)
        )

    if payment_status == "partial":
        return queryset.filter(
            payment_paid_sum__gt=0,
            payment_paid_sum__lt=F("amount")
        )

    if payment_status == "paid":
        return queryset.filter(
            payment_paid_sum=F("amount")
        )

    if payment_status == "overpaid":
        return queryset.filter(
            payment_paid_sum__gt=F("amount")
        )

    return queryset


def apply_ocr_status_filter(queryset, ocr_status):
    if ocr_status == "verified":
        return queryset.filter(
            amount_verified=True
        )

    if ocr_status == "unverified":
        return queryset.filter(
            amount_verified=False
        )

    return queryset


def apply_positive_payment_balance_filter(queryset):
    queryset = queryset.annotate(
        payment_paid_sum=Coalesce(
            Sum(
                "payments__amount",
                filter=Q(
                    payments__status=InvoicePayment.STATUS_POSTED
                )
            ),
            Value(
                Decimal("0.00"),
                output_field=DecimalField(
                    max_digits=12,
                    decimal_places=2
                )
            )
        )
    )

    return queryset.filter(
        payment_paid_sum__lt=F("amount")
    )


@login_required
def payment_schedule(request):

    filter_type = request.GET.get(
        'filter',
        'all'
    )

    search_query = request.GET.get(
        'q',
        ''
    ).strip()

    selected_status = request.GET.get(
        'status',
        'payment'
    )

    selected_priority = request.GET.get(
        'priority',
        ''
    )

    schedule_payment_status_filter = request.GET.get(
        'payment_status',
        ''
    )

    date_from = request.GET.get(
        'date_from',
        ''
    )

    date_to = request.GET.get(
        'date_to',
        ''
    )

    parsed_date_from = parse_date(date_from) if date_from else None

    parsed_date_to = parse_date(date_to) if date_to else None

    payment_statuses = [
        Invoice.STATUS_NEW,
        Invoice.STATUS_REVIEW,
        Invoice.STATUS_APPROVED,
    ]

    base_invoices = (
        Invoice.objects
        .select_related(
            'counterparty',
            'user'
        )
        .filter(
            status__in=payment_statuses
        )
    )

    today = date.today()

    week_end = today + timedelta(
        days=7
    )

    month_end = today + timedelta(
        days=30
    )

    total_count = base_invoices.count()

    today_count = base_invoices.filter(
        planned_payment_date=today
    ).count()

    week_count = base_invoices.filter(
        planned_payment_date__gte=today,
        planned_payment_date__lte=week_end
    ).count()

    month_count = base_invoices.filter(
        planned_payment_date__gte=today,
        planned_payment_date__lte=month_end
    ).count()

    overdue_count = base_invoices.filter(
        planned_payment_date__lt=today
    ).count()

    no_date_count = base_invoices.filter(
        planned_payment_date__isnull=True
    ).count()

    total_amount = (
        base_invoices.aggregate(
            total=Sum(
                'amount'
            )
        ).get(
            'total'
        )
        or 0
    )

    invoices = base_invoices

    if filter_type == 'today':

        invoices = invoices.filter(
            planned_payment_date=today
        )

    elif filter_type == 'week':

        invoices = invoices.filter(
            planned_payment_date__gte=today,
            planned_payment_date__lte=week_end
        )

    elif filter_type == 'month':

        invoices = invoices.filter(
            planned_payment_date__gte=today,
            planned_payment_date__lte=month_end
        )

    elif filter_type == 'overdue':

        invoices = invoices.filter(
            planned_payment_date__lt=today
        )

    elif filter_type == 'no_date':

        invoices = invoices.filter(
            planned_payment_date__isnull=True
        )

    if selected_status and selected_status not in [
        'payment',
        'all',
    ]:

        invoices = invoices.filter(
            status=selected_status
        )

    if selected_priority:

        invoices = invoices.filter(
            payment_priority=selected_priority
        )

    invoices = apply_payment_status_filter(
        invoices,
        schedule_payment_status_filter
    )

    if parsed_date_from and filter_type != 'no_date':

        invoices = invoices.filter(
            planned_payment_date__gte=parsed_date_from
        )

    if parsed_date_to and filter_type != 'no_date':

        invoices = invoices.filter(
            planned_payment_date__lte=parsed_date_to
        )

    if search_query:

        invoices = invoices.filter(
            Q(invoice_number__icontains=search_query)
            |
            Q(vendor__icontains=search_query)
            |
            Q(counterparty__name__icontains=search_query)
            |
            Q(original_filename__icontains=search_query)
            |
            Q(title__icontains=search_query)
            |
            Q(description__icontains=search_query)
        )

    filtered_count = invoices.count()

    filtered_amount = (
        invoices.aggregate(
            total=Sum(
                'amount'
            )
        ).get(
            'total'
        )
        or 0
    )

    priority_field = Invoice._meta.get_field(
        'payment_priority'
    )

    priority_choices = list(
        priority_field.choices or []
    )

    if not priority_choices:

        priority_choices = [
            (
                item,
                item
            )
            for item in (
                base_invoices
                .exclude(
                    payment_priority__isnull=True
                )
                .order_by(
                    '-payment_priority'
                )
                .values_list(
                    'payment_priority',
                    flat=True
                )
                .distinct()
            )
        ]

    invoices = apply_positive_payment_balance_filter(
        invoices
    )

    invoices = invoices.order_by(
        'planned_payment_date',
        '-payment_priority',
        'counterparty__name',
        'id'
    )

    return render(
        request,
        'invoices/payment_schedule.html',
        {
            'invoices': invoices,
            'today': today,
            'week_end': week_end,
            'month_end': month_end,
            'filter_type': filter_type,
            'search_query': search_query,
            'selected_status': selected_status,
            'selected_priority': selected_priority,
            'schedule_payment_status_filter': schedule_payment_status_filter,
            'payment_status_choices': PAYMENT_STATUS_FILTER_CHOICES,
            'date_from': date_from,
            'date_to': date_to,
            'status_choices': Invoice.STATUS_CHOICES,
            'priority_choices': priority_choices,
            'total_count': total_count,
            'today_count': today_count,
            'week_count': week_count,
            'month_count': month_count,
            'overdue_count': overdue_count,
            'no_date_count': no_date_count,
            'total_amount': total_amount,
            'filtered_count': filtered_count,
            'filtered_amount': filtered_amount,
        }
    )

@login_required
@require_payment_registry_permission(
    user_can_manage_payment_registry,
    'Нет прав на добавление счетов в реестр оплаты.',
)
def add_to_payment_registry(request):

    if request.method != 'POST':

        messages.warning(
            request,
            'Добавлять счета в реестр можно только из формы.'
        )

        return redirect(
            'payment_schedule'
        )

    invoice_ids = request.POST.getlist(
        'invoice_ids'
    )

    if not invoice_ids:

        messages.warning(
            request,
            'Выбери хотя бы один счет для добавления в реестр.'
        )

        return redirect(
            'payment_schedule'
        )

    from ..payment_registry_services import (
        add_invoice_to_payment_registry,
        get_or_create_draft_payment_registry,
    )

    registry, created = get_or_create_draft_payment_registry(
        request.user
    )

    invoices = (
        Invoice.objects
        .select_related(
            'counterparty',
            'user'
        )
        .filter(
            id__in=invoice_ids
        )
    )

    added_count = 0
    skipped_messages = []
    warning_messages = []

    for invoice in invoices:

        item, errors, warnings = add_invoice_to_payment_registry(
            invoice,
            registry
        )

        if item:

            added_count += 1

        if errors:

            skipped_messages.append(
                f'#{invoice.id}: ' + '; '.join(errors)
            )

        if warnings:

            warning_messages.append(
                f'#{invoice.id}: ' + '; '.join(warnings)
            )

    if added_count:

        messages.success(
            request,
            f'Добавлено счетов в реестр №{registry.id}: {added_count}.'
        )

    if skipped_messages:

        messages.warning(
            request,
            'Не добавлено: ' + ' | '.join(skipped_messages[:5])
        )

    if warning_messages:

        messages.info(
            request,
            'Предупреждения: ' + ' | '.join(warning_messages[:5])
        )

    if created and not added_count:

        registry.delete()

    return redirect(
        'payment_registry'
    )

@login_required
@require_payment_registry_permission(
    user_can_manage_payment_registry,
    'Нет прав на удаление счетов из черновика реестра.',
)
def remove_from_payment_registry_item(request, item_id):

    if request.method != 'POST':

        messages.warning(
            request,
            'Удалять счета из черновика можно только из формы.'
        )

        return redirect(
            'payment_registry'
        )

    from ..models import PaymentRegistry, PaymentRegistryItem
    from ..payment_registry_services import recalculate_payment_registry

    item = (
        PaymentRegistryItem.objects
        .select_related(
            'registry',
            'invoice',
        )
        .filter(
            id=item_id,
            registry__status=PaymentRegistry.STATUS_DRAFT,
            registry__created_by=request.user,
        )
        .exclude(
            status=PaymentRegistryItem.STATUS_CANCELLED
        )
        .first()
    )

    if not item:

        messages.warning(
            request,
            'Строка реестра не найдена или уже удалена.'
        )

        return redirect(
            'payment_registry'
        )

    registry = item.registry
    invoice_id = item.invoice_id

    item.status = PaymentRegistryItem.STATUS_CANCELLED
    item.save(
        update_fields=(
            'status',
        )
    )

    recalculate_payment_registry(
        registry
    )

    messages.success(
        request,
        f'Счёт #{invoice_id} удалён из черновика реестра №{registry.id}.'
    )

    return redirect(
        'payment_registry'
    )

@login_required
@require_payment_registry_permission(
    user_can_check_payment_registry,
    'Нет прав на проверку реестра оплаты.',
)
def check_payment_registry_view(request, registry_id):

    if request.method != 'POST':

        messages.warning(
            request,
            'Проверять реестр можно только из формы.'
        )

        return redirect(
            'payment_registry'
        )

    from ..models import PaymentRegistry
    from ..payment_registry_services import check_payment_registry

    registry = (
        PaymentRegistry.objects
        .filter(
            id=registry_id,
            created_by=request.user,
            status=PaymentRegistry.STATUS_DRAFT,
        )
        .first()
    )

    if not registry:

        messages.warning(
            request,
            'Черновик реестра не найден.'
        )

        return redirect(
            'payment_registry'
        )

    result = check_payment_registry(
        registry
    )

    if result['items_count'] == 0:

        messages.warning(
            request,
            f'Реестр №{registry.id} пуст. Сначала добавь счета.'
        )

        return redirect(
            'payment_registry'
        )

    if result['errors_count']:

        messages.warning(
            request,
            f'Реестр №{registry.id} не готов к выгрузке: ошибок {result["errors_count"]}.'
        )

        for error in result['errors'][:5]:

            messages.warning(
                request,
                f'Счёт #{error["invoice_id"]}: ' + '; '.join(error['messages'])
            )

    else:

        messages.success(
            request,
            f'Реестр №{registry.id} проверен: к выгрузке готово {result["ready_count"]} счетов.'
        )

    if result['warnings_count']:

        messages.info(
            request,
            f'Предупреждений: {result["warnings_count"]}.'
        )

    return redirect(
        'payment_registry'
    )

@login_required
@require_payment_registry_permission(
    user_can_export_payment_registry,
    'Нет прав на выгрузку реестра оплаты.',
)
def export_payment_registry_draft_excel(request, registry_id):

    if request.method != 'POST':

        messages.warning(
            request,
            'Выгрузка реестра выполняется только из формы.'
        )

        return redirect(
            'payment_registry'
        )

    from decimal import Decimal
    from io import BytesIO

    from django.http import HttpResponse
    from django.utils import timezone

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    from ..models import PaymentRegistry, PaymentRegistryItem
    from ..payment_registry_services import (
        check_payment_registry,
        recalculate_payment_registry,
    )

    registry = (
        PaymentRegistry.objects
        .filter(
            id=registry_id,
            created_by=request.user,
            status=PaymentRegistry.STATUS_DRAFT,
        )
        .first()
    )

    if not registry:

        messages.warning(
            request,
            'Черновик реестра не найден или уже выгружен.'
        )

        return redirect(
            'payment_registry'
        )

    check_result = check_payment_registry(
        registry
    )

    if check_result['items_count'] == 0:

        messages.warning(
            request,
            f'Реестр №{registry.id} пуст. Сначала добавь счета.'
        )

        return redirect(
            'payment_registry'
        )

    if check_result['errors_count']:

        messages.warning(
            request,
            f'Реестр №{registry.id} нельзя выгрузить: ошибок {check_result["errors_count"]}.'
        )

        return redirect(
            'payment_registry'
        )

    items = (
        registry.items
        .select_related(
            'invoice',
            'invoice__counterparty',
            'invoice__user',
        )
        .exclude(
            status=PaymentRegistryItem.STATUS_CANCELLED
        )
        .order_by(
            'planned_payment_date',
            'invoice_id',
        )
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f'Registry {registry.id}'

    headers = [
        '№ реестра',
        'ID счета',
        'Номер счета',
        'Контрагент',
        'ИНН',
        'КПП',
        'Банк',
        'Расчетный счет',
        'БИК',
        'Корр. счет',
        'Сумма',
        'Дата оплаты',
        'Назначение платежа',
        'Ответственный',
    ]

    ws.append(headers)

    header_fill = PatternFill(
        fill_type='solid',
        fgColor='1F2937',
    )

    for cell in ws[1]:
        cell.font = Font(
            bold=True,
            color='FFFFFF',
        )
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal='center',
            vertical='center',
        )

    for item in items:

        invoice = item.invoice
        counterparty = invoice.counterparty

        amount = item.amount or Decimal('0')
        payment_date = item.planned_payment_date or invoice.planned_payment_date

        purpose = (
            getattr(invoice, 'payment_purpose', '')
            or getattr(invoice, 'purpose', '')
            or getattr(invoice, 'description', '')
            or ''
        )

        ws.append(
            [
                registry.id,
                invoice.id,
                invoice.invoice_number or '',
                counterparty.name if counterparty else '',
                getattr(counterparty, 'inn', '') if counterparty else '',
                getattr(counterparty, 'kpp', '') if counterparty else '',
                getattr(counterparty, 'bank_name', '') if counterparty else '',
                getattr(counterparty, 'account_number', '') if counterparty else '',
                getattr(counterparty, 'bik', '') if counterparty else '',
                getattr(counterparty, 'correspondent_account', '') if counterparty else '',
                float(amount),
                payment_date.strftime('%d.%m.%Y') if payment_date else '',
                purpose,
                invoice.user.get_username() if invoice.user else '',
            ]
        )

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:
            value = str(cell.value or '')
            max_length = max(
                max_length,
                len(value),
            )

        ws.column_dimensions[column_letter].width = min(
            max_length + 2,
            42,
        )

    ws.freeze_panes = 'A2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    now = timezone.now()

    registry.status = PaymentRegistry.STATUS_EXPORTED
    registry.exported_by = request.user
    registry.exported_at = now
    registry.save(
        update_fields=(
            'status',
            'exported_by',
            'exported_at',
        )
    )

    items.update(
        status=PaymentRegistryItem.STATUS_EXPORTED,
        exported_at=now,
    )

    recalculate_payment_registry(
        registry
    )

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    response['Content-Disposition'] = (
        f'attachment; filename="payment_registry_{registry.id}.xlsx"'
    )

    return response

@login_required
@require_payment_registry_permission(
    user_can_export_payment_registry,
    'Нет прав на выгрузку реестра оплаты.',
)
def export_payment_registry_draft_1c(request, registry_id):

    if request.method != 'POST':

        messages.warning(
            request,
            'Выгрузка реестра выполняется только из формы.'
        )

        return redirect(
            'payment_registry'
        )

    import csv
    from io import StringIO

    from django.http import HttpResponse
    from django.utils import timezone

    from ..models import PaymentRegistry, PaymentRegistryItem
    from ..payment_registry_services import (
        check_payment_registry,
        recalculate_payment_registry,
    )

    registry = (
        PaymentRegistry.objects
        .filter(
            id=registry_id,
            created_by=request.user,
            status=PaymentRegistry.STATUS_DRAFT,
        )
        .first()
    )

    if not registry:

        messages.warning(
            request,
            'Черновик реестра не найден или уже выгружен.'
        )

        return redirect(
            'payment_registry'
        )

    check_result = check_payment_registry(
        registry
    )

    if check_result['items_count'] == 0:

        messages.warning(
            request,
            f'Реестр №{registry.id} пуст. Сначала добавь счета.'
        )

        return redirect(
            'payment_registry'
        )

    if check_result['errors_count']:

        messages.warning(
            request,
            f'Реестр №{registry.id} нельзя выгрузить: ошибок {check_result["errors_count"]}.'
        )

        return redirect(
            'payment_registry'
        )

    items = (
        registry.items
        .select_related(
            'invoice',
            'invoice__counterparty',
            'invoice__user',
        )
        .exclude(
            status=PaymentRegistryItem.STATUS_CANCELLED
        )
        .order_by(
            'planned_payment_date',
            'invoice_id',
        )
    )

    buffer = StringIO()
    writer = csv.writer(
        buffer,
        delimiter=';',
        lineterminator='\n',
    )

    writer.writerow(
        [
            'RegistryID',
            'InvoiceID',
            'InvoiceNumber',
            'Counterparty',
            'INN',
            'KPP',
            'Bank',
            'BankAccount',
            'BIK',
            'CorrAccount',
            'Amount',
            'PaymentDate',
            'Purpose',
        ]
    )

    for item in items:

        invoice = item.invoice
        counterparty = invoice.counterparty
        payment_date = item.planned_payment_date or invoice.planned_payment_date

        purpose = (
            getattr(invoice, 'payment_purpose', '')
            or getattr(invoice, 'purpose', '')
            or getattr(invoice, 'description', '')
            or ''
        )

        writer.writerow(
            [
                registry.id,
                invoice.id,
                invoice.invoice_number or '',
                counterparty.name if counterparty else '',
                getattr(counterparty, 'inn', '') if counterparty else '',
                getattr(counterparty, 'kpp', '') if counterparty else '',
                getattr(counterparty, 'bank_name', '') if counterparty else '',
                getattr(counterparty, 'account_number', '') if counterparty else '',
                getattr(counterparty, 'bik', '') if counterparty else '',
                getattr(counterparty, 'correspondent_account', '') if counterparty else '',
                str(item.amount).replace('.', ','),
                payment_date.strftime('%d.%m.%Y') if payment_date else '',
                purpose,
            ]
        )

    now = timezone.now()

    registry.status = PaymentRegistry.STATUS_EXPORTED
    registry.exported_by = request.user
    registry.exported_at = now
    registry.save(
        update_fields=(
            'status',
            'exported_by',
            'exported_at',
        )
    )

    items.update(
        status=PaymentRegistryItem.STATUS_EXPORTED,
        exported_at=now,
    )

    recalculate_payment_registry(
        registry
    )

    content = '\ufeff' + buffer.getvalue()

    response = HttpResponse(
        content,
        content_type='text/plain; charset=utf-8',
    )

    response['Content-Disposition'] = (
        f'attachment; filename="payment_registry_{registry.id}_1c.txt"'
    )

    return response

@login_required
@require_payment_registry_permission(
    user_can_mark_payment_registry_paid,
    'Нет прав на отметку реестра оплаченным.',
)
def mark_payment_registry_paid(request, registry_id):
    registry = get_object_or_404(
        PaymentRegistry,
        id=registry_id
    )

    if (
        not request.user.is_staff
        and not request.user.is_superuser
        and registry.created_by_id != request.user.id
    ):
        raise PermissionDenied

    if request.method != "POST":
        return redirect(
            "payment_registry_detail",
            registry_id=registry.id
        )

    try:
        result = mark_payment_registry_as_paid(
            registry,
            user=request.user
        )
    except ValueError as error:
        messages.error(
            request,
            str(error)
        )

        return redirect(
            "payment_registry_detail",
            registry_id=registry.id
        )

    messages.success(
        request,
        (
            "Реестр отмечен оплаченным. "
            f"Создано оплат: {result.get('paid_count', 0)}. "
            f"Пропущено закрытых счетов: {result.get('skipped_count', 0)}."
        )
    )

    return redirect(
        "payment_registry_detail",
        registry_id=registry.id
    )

@login_required
@require_payment_registry_permission(
    user_can_cancel_payment_registry,
    'Нет прав на отмену реестра оплаты.',
)
def cancel_payment_registry_view(request, registry_id):

    if request.method != 'POST':

        messages.warning(
            request,
            'Отменить реестр можно только из формы.'
        )

        return redirect(
            'payment_registry_detail',
            registry_id=registry_id,
        )

    from ..models import PaymentRegistry
    from ..payment_registry_services import cancel_payment_registry

    registry = (
        PaymentRegistry.objects
        .filter(
            id=registry_id,
        )
        .first()
    )

    if not registry:

        messages.warning(
            request,
            'Реестр оплаты не найден.'
        )

        return redirect(
            'payment_registry_history'
        )

    if not request.user.is_staff and registry.created_by_id != request.user.id:

        messages.warning(
            request,
            'Нет доступа к этому реестру.'
        )

        return redirect(
            'payment_registry_history'
        )

    reason = request.POST.get(
        'reason',
        ''
    ).strip()

    cancelled = cancel_payment_registry(
        registry,
        user=request.user,
        reason=reason,
    )

    if not cancelled:

        messages.warning(
            request,
            'Можно отменить только черновик или проверенный реестр.'
        )

        return redirect(
            'payment_registry_detail',
            registry_id=registry.id,
        )

    messages.success(
        request,
        f'Реестр оплаты №{registry.id} отменён.'
    )

    return redirect(
        'payment_registry_detail',
        registry_id=registry.id,
    )

@login_required
def payment_registry_detail(request, registry_id):

    from ..models import PaymentRegistry, PaymentRegistryItem
    from ..payment_registry_services import check_payment_registry

    registry = (
        PaymentRegistry.objects
        .select_related(
            'created_by',
            'checked_by',
            'exported_by',
        )
        .filter(
            id=registry_id,
        )
        .first()
    )

    if not registry:

        messages.warning(
            request,
            'Реестр оплаты не найден.'
        )

        return redirect(
            'payment_registry_history'
        )

    if not request.user.is_staff and registry.created_by_id != request.user.id:

        messages.warning(
            request,
            'Нет доступа к этому реестру.'
        )

        return redirect(
            'payment_registry_history'
        )

    registry_items = (
        registry.items
        .select_related(
            'invoice',
            'invoice__counterparty',
            'invoice__user',
        )
        .exclude(
            status=PaymentRegistryItem.STATUS_CANCELLED
        )
        .order_by(
            'planned_payment_date',
            'invoice_id',
        )
    )

    check_result = None

    if registry.status == PaymentRegistry.STATUS_DRAFT:

        check_result = check_payment_registry(
            registry
        )

    return render(
        request,
        'invoices/payment_registry_detail.html',
        {
            'page_title': f'Реестр оплаты №{registry.id}',
            'registry': registry,
            'registry_items': registry_items,
            'check_result': check_result,
        }
    )

@login_required
def payment_registry_history(request):

    from django.core.paginator import Paginator
    from django.db.models import Sum, Q

    from ..models import PaymentRegistry

    status_filter = request.GET.get(
        'status',
        ''
    ).strip()

    search_query = request.GET.get(
        'q',
        ''
    ).strip()

    registries = (
        PaymentRegistry.objects
        .select_related(
            'created_by',
            'checked_by',
            'exported_by',
        )
        .all()
        .order_by(
            '-created_at',
        )
    )

    if not request.user.is_staff:

        registries = registries.filter(
            created_by=request.user,
        )

    if status_filter:

        registries = registries.filter(
            status=status_filter,
        )

    if search_query:

        registries = registries.filter(
            Q(title__icontains=search_query)
            | Q(comment__icontains=search_query)
            | Q(created_by__username__icontains=search_query)
            | Q(exported_by__username__icontains=search_query)
        )

    total_registries = registries.count()

    total_amount = (
        registries.aggregate(
            total=Sum('total_amount')
        ).get('total')
        or 0
    )

    draft_count = registries.filter(
        status=PaymentRegistry.STATUS_DRAFT,
    ).count()

    exported_count = registries.filter(
        status=PaymentRegistry.STATUS_EXPORTED,
    ).count()

    paid_count = registries.filter(
        status=PaymentRegistry.STATUS_PAID,
    ).count()

    paginator = Paginator(
        registries,
        20,
    )

    page_obj = paginator.get_page(
        request.GET.get('page')
    )

    return render(
        request,
        'invoices/payment_registry_history.html',
        {
            'page_title': 'История реестров оплаты',
            'page_obj': page_obj,
            'registries': page_obj.object_list,
            'status_filter': status_filter,
            'search_query': search_query,
            'status_choices': PaymentRegistry.STATUS_CHOICES,
            'total_registries': total_registries,
            'total_amount': total_amount,
            'draft_count': draft_count,
            'exported_count': exported_count,
            'paid_count': paid_count,
        }
    )

@login_required
def payment_registry(request):

    selected_status = request.GET.get(
        'status',
        Invoice.STATUS_APPROVED
    )

    selected_counterparty = request.GET.get(
        'counterparty',
        ''
    )

    registry_payment_status_filter = request.GET.get(
        'payment_status',
        ''
    )

    ocr_status_filter = request.GET.get(
        'ocr_status',
        ''
    )

    search_query = request.GET.get(
        'q',
        ''
    ).strip()

    date_from = request.GET.get(
        'date_from',
        ''
    ).strip()

    date_to = request.GET.get(
        'date_to',
        ''
    ).strip()

    from ..models import PaymentRegistry, PaymentRegistryItem
    from ..payment_registry_services import ACTIVE_REGISTRY_STATUSES, check_payment_registry

    from ..payment_registry_services import (
        get_or_create_draft_payment_registry,
    )

    draft_registry, draft_registry_created = get_or_create_draft_payment_registry(request.user)

    draft_registry_items = PaymentRegistryItem.objects.none()
    draft_registry_check_result = None

    if draft_registry:

        draft_registry_items = (
            draft_registry.items
            .select_related(
                'invoice',
                'invoice__counterparty',
                'invoice__user',
            )
            .exclude(
                status=PaymentRegistryItem.STATUS_CANCELLED
            )
            .order_by(
                'planned_payment_date',
                'invoice_id',
            )
        )

        draft_registry_check_result = check_payment_registry(
            draft_registry
        )

    active_registry_invoice_ids = (
        PaymentRegistryItem.objects
        .filter(
            registry__status__in=ACTIVE_REGISTRY_STATUSES,
        )
        .exclude(
            status=PaymentRegistryItem.STATUS_CANCELLED
        )
        .values_list(
            'invoice_id',
            flat=True,
        )
    )

    invoices = (
        Invoice.objects
        .select_related(
            'counterparty',
            'user'
        )
        .exclude(
            status=Invoice.STATUS_PAID
        )
        .exclude(
            id__in=active_registry_invoice_ids
        )
    )

    if selected_status and selected_status != 'all':

        invoices = invoices.filter(
            status=selected_status
        )

    if selected_counterparty:

        invoices = invoices.filter(
            counterparty_id=selected_counterparty
        )

    invoices = apply_payment_status_filter(
        invoices,
        registry_payment_status_filter
    )

    invoices = apply_ocr_status_filter(
        invoices,
        ocr_status_filter
    )

    if search_query:

        invoices = invoices.filter(
            Q(title__icontains=search_query)
            |
            Q(original_filename__icontains=search_query)
            |
            Q(invoice_number__icontains=search_query)
            |
            Q(vendor__icontains=search_query)
            |
            Q(counterparty__name__icontains=search_query)
        )

    if date_from:

        invoices = invoices.filter(
            planned_payment_date__gte=date_from
        )

    if date_to:

        invoices = invoices.filter(
            planned_payment_date__lte=date_to
        )

    invoices = apply_positive_payment_balance_filter(
        invoices
    )

    total_amount = (
        invoices.aggregate(
            total=Sum(
                'amount'
            )
        ).get(
            'total'
        )
        or 0
    )

    counterparties = (
        Counterparty.objects
        .filter(
            invoices__isnull=False
        )
        .distinct()
        .order_by(
            'name'
        )
    )

    invoices = invoices.order_by(
        'planned_payment_date',
        '-payment_priority',
        'counterparty__name',
        'id'
    )

    # OCR_REGISTRY_SUMMARY_CONTEXT_V3
    ocr_registry_draft_items = list(draft_registry_items or [])
    draft_registry_items = ocr_registry_draft_items

    ocr_registry_invoice_map = {}

    for item in ocr_registry_draft_items:
        if item.invoice_id:
            ocr_registry_invoice_map[item.invoice_id] = item.invoice

    for invoice in list(invoices or []):
        if invoice.id:
            ocr_registry_invoice_map[invoice.id] = invoice

    ocr_registry_invoices = list(ocr_registry_invoice_map.values())
    ocr_registry_items_count = len(ocr_registry_invoices)
    ocr_registry_ready_count = sum(
        1
        for invoice in ocr_registry_invoices
        if invoice.amount_verified
    )
    ocr_registry_errors_count = sum(
        1
        for invoice in ocr_registry_invoices
        if not invoice.amount_verified
    )

    return render(
        request,
        'invoices/payment_registry.html',
        {
            'invoices': invoices,
            'counterparties': counterparties,
            'total_amount': total_amount,
            'selected_status': selected_status,
            'selected_counterparty': selected_counterparty,
            'registry_payment_status_filter': registry_payment_status_filter,
            'payment_status_choices': PAYMENT_STATUS_FILTER_CHOICES,
            'ocr_status_filter': ocr_status_filter,
            'ocr_status_choices': OCR_STATUS_FILTER_CHOICES,
            'search_query': search_query,
            'date_from': date_from,
            'date_to': date_to,
            'status_choices': Invoice.STATUS_CHOICES,
            'draft_registry': draft_registry,
            'draft_registry_items': draft_registry_items,
            "ocr_registry_items_count": ocr_registry_items_count,
            "ocr_registry_ready_count": ocr_registry_ready_count,
            "ocr_registry_errors_count": ocr_registry_errors_count,
            'draft_registry_items_count': draft_registry.items_count if draft_registry else 0,
            'draft_registry_total_amount': draft_registry.total_amount if draft_registry else 0,
            'draft_registry_check_result': draft_registry_check_result,
        }
    )

@login_required
@require_payment_registry_permission(
    user_can_export_payment_registry,
    'Нет прав на выгрузку реестра оплаты.',
)
def export_payment_registry_excel(request):

    selected_status = request.GET.get(
        'status',
        Invoice.STATUS_APPROVED
    )

    selected_counterparty = request.GET.get(
        'counterparty',
        ''
    )

    search_query = request.GET.get(
        'q',
        ''
    ).strip()

    date_from = request.GET.get(
        'date_from',
        ''
    ).strip()

    date_to = request.GET.get(
        'date_to',
        ''
    ).strip()

    invoices = (
        Invoice.objects
        .select_related(
            'counterparty',
            'user'
        )
        .exclude(
            status=Invoice.STATUS_PAID
        )
    )

    if selected_status and selected_status != 'all':

        invoices = invoices.filter(
            status=selected_status
        )

    if selected_counterparty:

        invoices = invoices.filter(
            counterparty_id=selected_counterparty
        )

    if search_query:

        invoices = invoices.filter(
            Q(title__icontains=search_query)
            |
            Q(original_filename__icontains=search_query)
            |
            Q(invoice_number__icontains=search_query)
            |
            Q(vendor__icontains=search_query)
            |
            Q(counterparty__name__icontains=search_query)
        )

    if date_from:

        invoices = invoices.filter(
            planned_payment_date__gte=date_from
        )

    if date_to:

        invoices = invoices.filter(
            planned_payment_date__lte=date_to
        )

    invoices = invoices.order_by(
        'planned_payment_date',
        '-payment_priority',
        'counterparty__name',
        'id'
    )

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = 'Реестр оплаты'

    headers = [
        'ID',
        'Контрагент',
        'ИНН',
        'КПП',
        'Номер счета',
        'Дата счета',
        'Сумма',
        'Плановая дата оплаты',
        'Приоритет',
        'Статус',
        'Назначение платежа',
    ]

    sheet.append(
        headers
    )

    header_fill = PatternFill(
        fill_type='solid',
        fgColor='E5E7EB'
    )

    for cell in sheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.fill = header_fill

        cell.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )

    total_amount = Decimal(
        '0.00'
    )

    for invoice in invoices:

        if invoice.counterparty:

            counterparty_name = invoice.counterparty.name or ''
            inn = invoice.counterparty.inn or ''
            kpp = invoice.counterparty.kpp or ''

        else:

            counterparty_name = invoice.vendor or ''
            inn = ''
            kpp = ''

        amount = invoice.amount or Decimal(
            '0.00'
        )

        total_amount += amount

        if invoice.invoice_number:

            payment_purpose = (
                f'Оплата по счету №{invoice.invoice_number}'
            )

            if invoice.invoice_date:

                payment_purpose += (
                    f' от {invoice.invoice_date}'
                )

        else:

            payment_purpose = (
                f'Оплата по счету ID {invoice.id}'
            )

        sheet.append(
            [
                invoice.id,
                counterparty_name,
                inn,
                kpp,
                invoice.invoice_number or '',
                invoice.invoice_date or '',
                amount,
                (
                    invoice.planned_payment_date.strftime(
                        '%d.%m.%Y'
                    )
                    if invoice.planned_payment_date
                    else ''
                ),
                invoice.payment_priority,
                invoice.get_status_display(),
                payment_purpose,
            ]
        )

    total_row = sheet.max_row + 2

    sheet.cell(
        row=total_row,
        column=6,
        value='Итого:'
    )

    sheet.cell(
        row=total_row,
        column=7,
        value=total_amount
    )

    sheet.cell(
        row=total_row,
        column=6
    ).font = Font(
        bold=True
    )

    sheet.cell(
        row=total_row,
        column=7
    ).font = Font(
        bold=True
    )

    for row in sheet.iter_rows():

        for cell in row:

            cell.alignment = Alignment(
                vertical='top',
                wrap_text=True
            )

    column_widths = {
        'A': 8,
        'B': 36,
        'C': 16,
        'D': 14,
        'E': 18,
        'F': 16,
        'G': 16,
        'H': 20,
        'I': 12,
        'J': 18,
        'K': 48,
    }

    for column_letter, width in column_widths.items():

        sheet.column_dimensions[
            column_letter
        ].width = width

    for row_number in range(
        2,
        sheet.max_row + 1
    ):

        sheet.cell(
            row=row_number,
            column=7
        ).number_format = '#,##0.00'

    sheet.freeze_panes = 'A2'

    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet'
        )
    )

    response[
        'Content-Disposition'
    ] = (
        'attachment; filename="payment_registry.xlsx"'
    )

    workbook.save(
        response
    )

    return response

@login_required
@require_payment_registry_permission(
    user_can_export_payment_registry,
    'Нет прав на выгрузку реестра оплаты.',
)
def export_payment_registry_1c(request):

    company = CompanyRequisites.objects.first()

    if not company:

        messages.error(
            request,
            'Сначала заполните реквизиты организации в админке.'
        )

        return redirect(
            'payment_registry'
        )

    selected_status = request.GET.get(
        'status',
        Invoice.STATUS_APPROVED
    )

    selected_counterparty = request.GET.get(
        'counterparty',
        ''
    )

    search_query = request.GET.get(
        'q',
        ''
    ).strip()

    date_from = request.GET.get(
        'date_from',
        ''
    ).strip()

    date_to = request.GET.get(
        'date_to',
        ''
    ).strip()

    invoices = (
        Invoice.objects
        .select_related(
            'counterparty',
            'user'
        )
        .exclude(
            status=Invoice.STATUS_PAID
        )
    )

    if selected_status and selected_status != 'all':

        invoices = invoices.filter(
            status=selected_status
        )

    if selected_counterparty:

        invoices = invoices.filter(
            counterparty_id=selected_counterparty
        )

    if search_query:

        invoices = invoices.filter(
            Q(title__icontains=search_query)
            |
            Q(original_filename__icontains=search_query)
            |
            Q(invoice_number__icontains=search_query)
            |
            Q(vendor__icontains=search_query)
            |
            Q(counterparty__name__icontains=search_query)
        )

    if date_from:

        invoices = invoices.filter(
            planned_payment_date__gte=date_from
        )

    if date_to:

        invoices = invoices.filter(
            planned_payment_date__lte=date_to
        )

    invoices = invoices.order_by(
        'planned_payment_date',
        '-payment_priority',
        'counterparty__name',
        'id'
    )

    missing_requisites = []

    for invoice in invoices:

        counterparty = invoice.counterparty

        if not counterparty:

            missing_requisites.append(
                f'Счет #{invoice.id}: контрагент не найден'
            )

            continue

        required_values = [
            counterparty.inn,
            counterparty.bank_name,
            counterparty.bik,
            counterparty.account_number,
        ]

        if any(
            not value
            for value in required_values
        ):

            missing_requisites.append(
                f'Счет #{invoice.id}: {counterparty.name}'
            )

    if missing_requisites:

        messages.error(
            request,
            (
                'Выгрузка 1С TXT остановлена. '
                'Есть контрагенты без обязательных платежных реквизитов.'
            )
        )

        return redirect(
            'counterparties_missing_requisites'
        )

    def clean_value(value):

        if value is None:

            return ''

        value = str(value)

        value = value.replace(
            '\r',
            ' '
        )

        value = value.replace(
            '\n',
            ' '
        )

        value = value.strip()

        return value

    def format_date(value):

        if value:

            return value.strftime(
                '%d.%m.%Y'
            )

        return date.today().strftime(
            '%d.%m.%Y'
        )

    def format_amount(value):

        if not value:

            return '0.00'

        return f'{value:.2f}'

    created_at = datetime.now()

    lines = []

    lines.append(
        '1CClientBankExchange'
    )

    lines.append(
        'ВерсияФормата=1.03'
    )

    lines.append(
        'Кодировка=Windows'
    )

    lines.append(
        'Отправитель=Project Zarya'
    )

    lines.append(
        'Получатель=1С'
    )

    lines.append(
        f'ДатаСоздания={created_at.strftime("%d.%m.%Y")}'
    )

    lines.append(
        f'ВремяСоздания={created_at.strftime("%H:%M:%S")}'
    )

    lines.append(
        f'РасчСчет={clean_value(company.account_number)}'
    )

    for invoice in invoices:

        if not invoice.counterparty:

            continue

        counterparty = invoice.counterparty

        amount = invoice.amount or 0

        payment_date = invoice.planned_payment_date or date.today()

        if invoice.invoice_number:

            payment_purpose = (
                f'Оплата по счету №{invoice.invoice_number}'
            )

            if invoice.invoice_date:

                payment_purpose += (
                    f' от {invoice.invoice_date}'
                )

        else:

            payment_purpose = (
                f'Оплата по счету ID {invoice.id}'
            )

        lines.append(
            'СекцияДокумент=Платежное поручение'
        )

        lines.append(
            f'Номер={invoice.id}'
        )

        lines.append(
            f'Дата={format_date(payment_date)}'
        )

        lines.append(
            f'Сумма={format_amount(amount)}'
        )

        lines.append(
            f'Плательщик={clean_value(company.name)}'
        )

        lines.append(
            f'ПлательщикИНН={clean_value(company.inn)}'
        )

        lines.append(
            f'ПлательщикКПП={clean_value(company.kpp)}'
        )

        lines.append(
            f'ПлательщикСчет={clean_value(company.account_number)}'
        )

        lines.append(
            f'ПлательщикБанк1={clean_value(company.bank_name)}'
        )

        lines.append(
            f'ПлательщикБИК={clean_value(company.bik)}'
        )

        lines.append(
            f'ПлательщикКорсчет={clean_value(company.correspondent_account)}'
        )

        lines.append(
            f'Получатель={clean_value(counterparty.name)}'
        )

        lines.append(
            f'ПолучательИНН={clean_value(counterparty.inn)}'
        )

        lines.append(
            f'ПолучательКПП={clean_value(counterparty.kpp)}'
        )

        lines.append(
            f'ПолучательСчет={clean_value(counterparty.account_number)}'
        )

        lines.append(
            f'ПолучательБанк1={clean_value(counterparty.bank_name)}'
        )

        lines.append(
            f'ПолучательБИК={clean_value(counterparty.bik)}'
        )

        lines.append(
            f'ПолучательКорсчет={clean_value(counterparty.correspondent_account)}'
        )

        lines.append(
            'ВидПлатежа=Электронно'
        )

        lines.append(
            'ВидОплаты=01'
        )

        lines.append(
            'Очередность=5'
        )

        lines.append(
            f'НазначениеПлатежа={clean_value(payment_purpose)}'
        )

        lines.append(
            'КонецДокумента'
        )

    lines.append(
        'КонецФайла'
    )

    content = '\r\n'.join(
        lines
    )

    response = HttpResponse(
        content.encode(
            'cp1251',
            errors='replace'
        ),
        content_type='text/plain; charset=windows-1251'
    )

    response[
        'Content-Disposition'
    ] = (
        'attachment; filename="payment_registry_1c.txt"'
    )

    return response
