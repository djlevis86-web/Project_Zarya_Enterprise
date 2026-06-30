from django.utils.dateparse import parse_date
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
import hashlib
import traceback
import uuid
from django.conf import settings
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from ..comment_forms import InvoiceCommentForm
from ..comment_models import InvoiceComment
from ..counterparty_service import (
    extract_requisites_near_vendor,
    normalize_counterparty_name,
)
from ..forms import (
    CounterpartyImportForm,
    CounterpartyManualForm,
    InvoiceCounterpartyAssignForm,
    InvoiceEditForm,
    InvoiceForm,
)
from ..log_service import create_invoice_log
from ..models import (
    CompanyRequisites,
    Counterparty,
    Invoice,
    InvoiceUploadBatch,
    OCRJob,
)
from ..one_c_import_service import import_counterparties_from_file
from ..ocr_processing_service import (
    apply_ocr_identity_to_invoice,
    get_duplicate_invoice_by_ocr_identity,
    read_and_parse_invoice_file,
    run_invoice_ocr_processing,
)
from ..ocr_verification_service import (
    apply_ocr_amount_to_invoice,
    sync_invoice_amount_verification,
)
from ..payment_registry_services import get_active_registry_items_for_invoice
from audit.models import AuditLog
from audit.services import log_action
from ..payment_registry_permissions import (
    require_payment_registry_permission,
    user_can_cancel_payment_registry,
    user_can_check_payment_registry,
    user_can_export_payment_registry,
    user_can_manage_payment_registry,
    user_can_mark_payment_registry_paid,
)

from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..forms import CounterpartyManualForm
from ..models import Counterparty, Invoice




from .counterparty_unmatched_views import (
    export_unmatched_counterparties_excel,
    unmatched_counterparties,
)


@staff_member_required
def import_counterparties_1c(request):

    result = None

    if request.method == 'POST':

        form = CounterpartyImportForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            uploaded_file = form.cleaned_data[
                'file'
            ]

            import_dir = Path(
                settings.MEDIA_ROOT
            ) / 'imports_1c'

            import_dir.mkdir(
                parents=True,
                exist_ok=True
            )

            storage = FileSystemStorage(
                location=str(import_dir)
            )

            saved_name = storage.save(
                uploaded_file.name,
                uploaded_file
            )

            file_path = import_dir / saved_name

            result = import_counterparties_from_file(
                file_path=file_path,
                clear_ocr=form.cleaned_data[
                    'clear_ocr'
                ],
                deactivate_missing=form.cleaned_data[
                    'deactivate_missing'
                ]
            )

            messages.success(
                request,
                (
                    'Импорт справочника из 1С завершен. '
                    f'Создано: {result["created"]}, '
                    f'обновлено: {result["updated"]}, '
                    f'пропущено: {result["skipped"]}.'
                )
            )

    else:

        form = CounterpartyImportForm()

    return render(
        request,
        'invoices/import_counterparties_1c.html',
        {
            'form': form,
            'result': result,
        }
    )

@staff_member_required
def rematch_counterparties_1c(request):

    if request.method != 'POST':

        return redirect(
            'unmatched_counterparties'
        )

    from ..counterparty_service import get_or_create_counterparty_from_invoice

    invoices = Invoice.objects.all().order_by(
        'id'
    )

    matched = 0
    not_found = 0

    for invoice in invoices:

        invoice.counterparty = None

        counterparty = get_or_create_counterparty_from_invoice(
            invoice
        )

        invoice.counterparty = counterparty

        invoice.save(
            update_fields=[
                'counterparty',
                'counterparty_match_status',
                'counterparty_match_comment',
            ]
        )

        if counterparty:

            matched += 1

        else:

            not_found += 1

    messages.success(
        request,
        (
            'Пересопоставление завершено. '
            f'Найдено: {matched}, '
            f'не найдено: {not_found}.'
        )
    )

    return redirect(
        'unmatched_counterparties'
    )

@staff_member_required
def counterparties_missing_requisites(request):

    payment_statuses = [
        Invoice.STATUS_NEW,
        Invoice.STATUS_REVIEW,
        Invoice.STATUS_APPROVED,
    ]

    counterparties = (
        Counterparty.objects
        .filter(
            is_active=True,
            invoices__status__in=payment_statuses
        )
        .filter(
            Q(source=Counterparty.SOURCE_1C)
            |
            Q(source=Counterparty.SOURCE_MANUAL)
        )
        .filter(
            Q(inn__isnull=True)
            |
            Q(inn='')
            |
            Q(bank_name__isnull=True)
            |
            Q(bank_name='')
            |
            Q(bik__isnull=True)
            |
            Q(bik='')
            |
            Q(account_number__isnull=True)
            |
            Q(account_number='')
        )
        .annotate(
            invoices_count=Count(
                'invoices',
                filter=Q(
                    invoices__status__in=payment_statuses
                ),
                distinct=True
            ),
            invoices_total=Sum(
                'invoices__amount',
                filter=Q(
                    invoices__status__in=payment_statuses
                )
            )
        )
        .distinct()
        .order_by(
            'name'
        )
    )

    return render(
        request,
        'invoices/counterparties_missing_requisites.html',
        {
            'counterparties': counterparties,
            'counterparties_count': counterparties.count(),
        }
    )
