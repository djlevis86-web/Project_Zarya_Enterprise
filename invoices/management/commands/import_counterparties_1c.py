import csv
import re

from pathlib import Path

from django.core.management.base import BaseCommand
from django.utils import timezone

from openpyxl import load_workbook

from invoices.models import Counterparty


HEADER_ALIASES = {
    'external_id_1c': [
        'код',
        'код 1с',
        'ссылка',
        'guid',
        'идентификатор',
        'external_id_1c',
    ],
    'name': [
        'наименование',
        'краткое наименование',
        'контрагент',
        'name',
    ],
    'full_name': [
        'полное наименование',
        'полное имя',
        'full_name',
    ],
    'inn': [
        'инн',
    ],
    'kpp': [
        'кпп',
    ],
    'bank_name': [
        'банк',
        'наименование банка',
        'банк получателя',
    ],
    'bik': [
        'бик',
    ],
    'account_number': [
        'расчетный счет',
        'расчётный счет',
        'р/с',
        'рс',
        'счет',
        'счёт',
    ],
    'correspondent_account': [
        'корреспондентский счет',
        'корреспондентский счёт',
        'корр счет',
        'корр. счет',
        'к/с',
        'кс',
    ],
    'is_active': [
        'активен',
        'используется',
    ],
    'deletion_mark': [
        'пометка удаления',
        'удален',
        'удалён',
    ],
}


def normalize_header(value):

    value = str(value or '').strip().lower()

    value = value.replace(
        'ё',
        'е'
    )

    value = re.sub(
        r'[^a-zа-я0-9]+',
        '',
        value
    )

    return value


def clean_value(value):

    if value is None:

        return ''

    value = str(value).strip()

    if value.endswith(
        '.0'
    ):

        value = value[:-2]

    return value


def only_digits(value):

    value = clean_value(
        value
    )

    return re.sub(
        r'\D',
        '',
        value
    )


def get_mapped_value(row, field_name):

    aliases = HEADER_ALIASES.get(
        field_name,
        []
    )

    normalized_row = {
        normalize_header(key): value
        for key, value in row.items()
    }

    for alias in aliases:

        alias_key = normalize_header(
            alias
        )

        if alias_key in normalized_row:

            return clean_value(
                normalized_row[alias_key]
            )

    return ''


def parse_bool(value, default=True):

    value = clean_value(
        value
    ).lower()

    if value in [
        '',
        'none',
        'null',
    ]:

        return default

    if value in [
        'нет',
        'false',
        '0',
        'ложь',
        'неактивен',
    ]:

        return False

    if value in [
        'да',
        'true',
        '1',
        'истина',
        'активен',
    ]:

        return True

    return default


def parse_deletion_mark(value):

    value = clean_value(
        value
    ).lower()

    if value in [
        'да',
        'true',
        '1',
        'истина',
        'помечен',
        'пометка удаления',
    ]:

        return True

    return False


def read_xlsx(file_path, sheet_name=None):

    workbook = load_workbook(
        file_path,
        data_only=True
    )

    if sheet_name:

        sheet = workbook[
            sheet_name
        ]

    else:

        sheet = workbook.active

    rows = list(
        sheet.iter_rows(
            values_only=True
        )
    )

    if not rows:

        return []

    headers = [
        clean_value(value)
        for value in rows[0]
    ]

    result = []

    for raw_row in rows[1:]:

        row = {}

        for index, header in enumerate(headers):

            if not header:

                continue

            row[header] = (
                raw_row[index]
                if index < len(raw_row)
                else ''
            )

        if any(
            clean_value(value)
            for value in row.values()
        ):

            result.append(
                row
            )

    return result


def read_csv(file_path):

    encodings = [
        'utf-8-sig',
        'cp1251',
    ]

    for encoding in encodings:

        try:

            with open(
                file_path,
                'r',
                encoding=encoding,
                newline=''
            ) as file:

                sample = file.read(
                    4096
                )

                file.seek(
                    0
                )

                try:

                    dialect = csv.Sniffer().sniff(
                        sample,
                        delimiters=';,'
                    )

                except Exception:

                    dialect = csv.excel
                    dialect.delimiter = ';'

                reader = csv.DictReader(
                    file,
                    dialect=dialect
                )

                return list(
                    reader
                )

        except UnicodeDecodeError:

            continue

    raise UnicodeDecodeError(
        'csv',
        b'',
        0,
        1,
        'Не удалось определить кодировку CSV'
    )


class Command(BaseCommand):

    help = 'Импортирует справочник контрагентов из выгрузки 1С XLSX/CSV'

    def add_arguments(self, parser):

        parser.add_argument(
            'file_path',
            type=str,
            help='Путь к XLSX или CSV файлу из 1С'
        )

        parser.add_argument(
            '--sheet',
            type=str,
            default=None,
            help='Название листа XLSX'
        )

        parser.add_argument(
            '--clear-ocr',
            action='store_true',
            help='Удалить контрагентов, созданных OCR'
        )

        parser.add_argument(
            '--deactivate-missing',
            action='store_true',
            help='Деактивировать контрагентов 1С, которых нет в текущей выгрузке'
        )

    def handle(self, *args, **options):

        file_path = Path(
            options['file_path']
        )

        if not file_path.exists():

            self.stdout.write(
                self.style.ERROR(
                    f'Файл не найден: {file_path}'
                )
            )

            return

        if options['clear_ocr']:

            deleted_count, _ = Counterparty.objects.filter(
                source=Counterparty.SOURCE_OCR
            ).delete()

            self.stdout.write(
                self.style.WARNING(
                    f'Удалено OCR-контрагентов: {deleted_count}'
                )
            )

        extension = file_path.suffix.lower()

        if extension in [
            '.xlsx',
            '.xlsm',
        ]:

            rows = read_xlsx(
                file_path,
                sheet_name=options['sheet']
            )

        elif extension == '.csv':

            rows = read_csv(
                file_path
            )

        else:

            self.stdout.write(
                self.style.ERROR(
                    'Поддерживаются только .xlsx, .xlsm, .csv'
                )
            )

            return

        self.stdout.write(
            f'Строк в файле: {len(rows)}'
        )

        created = 0
        updated = 0
        skipped = 0

        imported_external_ids = []

        for row in rows:

            external_id_1c = get_mapped_value(
                row,
                'external_id_1c'
            )

            name = get_mapped_value(
                row,
                'name'
            )

            full_name = get_mapped_value(
                row,
                'full_name'
            )

            inn = only_digits(
                get_mapped_value(
                    row,
                    'inn'
                )
            )

            kpp = only_digits(
                get_mapped_value(
                    row,
                    'kpp'
                )
            )

            bank_name = get_mapped_value(
                row,
                'bank_name'
            )

            bik = only_digits(
                get_mapped_value(
                    row,
                    'bik'
                )
            )

            account_number = only_digits(
                get_mapped_value(
                    row,
                    'account_number'
                )
            )

            correspondent_account = only_digits(
                get_mapped_value(
                    row,
                    'correspondent_account'
                )
            )

            is_active = parse_bool(
                get_mapped_value(
                    row,
                    'is_active'
                ),
                default=True
            )

            if parse_deletion_mark(
                get_mapped_value(
                    row,
                    'deletion_mark'
                )
            ):

                is_active = False

            if not name and full_name:

                name = full_name

            if not name:

                skipped += 1

                continue

            lookup = None

            counterparty = None

            if external_id_1c:

                lookup = {
                    'external_id_1c': external_id_1c
                }

                imported_external_ids.append(
                    external_id_1c
                )

                counterparty = Counterparty.objects.filter(
                    **lookup
                ).first()

            if not counterparty and inn and kpp:

                counterparty = Counterparty.objects.filter(
                    inn=inn,
                    kpp=kpp,
                    source=Counterparty.SOURCE_1C
                ).first()

            if not counterparty and inn:

                counterparty = Counterparty.objects.filter(
                    inn=inn,
                    source=Counterparty.SOURCE_1C
                ).first()

            if not counterparty:

                counterparty = Counterparty.objects.filter(
                    name__iexact=name,
                    source=Counterparty.SOURCE_1C
                ).first()

            if counterparty:

                was_created = False

            else:

                counterparty = Counterparty()

                was_created = True

            counterparty.external_id_1c = external_id_1c or counterparty.external_id_1c
            counterparty.name = name
            counterparty.full_name = full_name or counterparty.full_name
            counterparty.inn = inn or counterparty.inn
            counterparty.kpp = kpp or counterparty.kpp
            counterparty.bank_name = bank_name or counterparty.bank_name
            counterparty.bik = bik or counterparty.bik
            counterparty.account_number = account_number or counterparty.account_number
            counterparty.correspondent_account = correspondent_account or counterparty.correspondent_account
            counterparty.source = Counterparty.SOURCE_1C
            counterparty.is_active = is_active
            counterparty.synced_at = timezone.now()
            counterparty.sync_comment = 'Импортировано из 1С'

            counterparty.save()

            if was_created:

                created += 1

            else:

                updated += 1

        if options['deactivate_missing'] and imported_external_ids:

            deactivated = Counterparty.objects.filter(
                source=Counterparty.SOURCE_1C
            ).exclude(
                external_id_1c__in=imported_external_ids
            ).update(
                is_active=False
            )

            self.stdout.write(
                self.style.WARNING(
                    f'Деактивировано отсутствующих в выгрузке: {deactivated}'
                )
            )

        self.stdout.write(
            self.style.SUCCESS(
                f'Готово. Создано: {created}, обновлено: {updated}, пропущено: {skipped}'
            )
        )

        self.stdout.write(
            f'Всего контрагентов 1С: {Counterparty.objects.filter(source=Counterparty.SOURCE_1C).count()}'
        )