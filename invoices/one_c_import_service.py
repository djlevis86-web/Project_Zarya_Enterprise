import csv
import re

from pathlib import Path

from django.utils import timezone

from openpyxl import load_workbook

from invoices.models import Counterparty


HEADER_ALIASES = {
    'external_id_1c': [
        'код',
        'код 1с',
        'ссылка',
        'guid',
        'идентификатор',
        'external_id_1c',
    ],
    'name': [
        'наименование',
        'краткое наименование',
        'контрагент',
        'name',
    ],
    'full_name': [
        'полное наименование',
        'полное имя',
        'full_name',
    ],
    'inn': [
        'инн',
    ],
    'kpp': [
        'кпп',
    ],
    'bank_name': [
        'банк',
        'наименование банка',
        'банк получателя',
    ],
    'bik': [
        'бик',
    ],
    'account_number': [
        'расчетный счет',
        'расчётный счет',
        'р/с',
        'рс',
        'счет',
        'счёт',
    ],
    'correspondent_account': [
        'корреспондентский счет',
        'корреспондентский счёт',
        'корр счет',
        'корр. счет',
        'к/с',
        'кс',
    ],
    'is_active': [
        'активен',
        'используется',
    ],
    'deletion_mark': [
        'пометка удаления',
        'удален',
        'удалён',
    ],
}


def normalize_header(value):

    value = str(value or '').strip().lower()

    value = value.replace(
        'ё',
        'е'
    )

    value = re.sub(
        r'[^a-zа-я0-9]+',
        '',
        value
    )

    return value


def clean_value(value):

    if value is None:

        return ''

    value = str(value).strip()

    if value.endswith(
        '.0'
    ):

        value = value[:-2]

    return value


def only_digits(value):

    value = clean_value(
        value
    )

    return re.sub(
        r'\D',
        '',
        value
    )


def get_mapped_value(row, field_name):

    aliases = HEADER_ALIASES.get(
        field_name,
        []
    )

    normalized_row = {
        normalize_header(key): value
        for key, value in row.items()
    }

    for alias in aliases:

        alias_key = normalize_header(
            alias
        )

        if alias_key in normalized_row:

            return clean_value(
                normalized_row[alias_key]
            )

    return ''


def parse_bool(value, default=True):

    value = clean_value(
        value
    ).lower()

    if value in [
        '',
        'none',
        'null',
    ]:

        return default

    if value in [
        'нет',
        'false',
        '0',
        'ложь',
        'неактивен',
    ]:

        return False

    if value in [
        'да',
        'true',
        '1',
        'истина',
        'активен',
    ]:

        return True

    return default


def parse_deletion_mark(value):

    value = clean_value(
        value
    ).lower()

    if value in [
        'да',
        'true',
        '1',
        'истина',
        'помечен',
        'пометка удаления',
    ]:

        return True

    return False


def read_xlsx(file_path, sheet_name=None):
    """
    Читает XLSX/XLSM выгрузку 1С.

    Особенность выгрузок 1С:
    первые строки часто служебные:
    - Параметры;
    - Тип объекта;
    - Имя объекта;
    - Имя таблицы.

    Поэтому строку заголовков ищем автоматически среди первых 40 строк.
    """

    workbook = load_workbook(
        file_path,
        data_only=True,
        read_only=True
    )

    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise Exception(
                    f'Лист "{sheet_name}" не найден. '
                    f'Доступные листы: {", ".join(workbook.sheetnames)}'
                )

            worksheet = workbook[sheet_name]

        else:
            worksheet = workbook.active

        def cell_to_text(value):
            if value is None:
                return ''

            text = str(value)
            text = text.replace('\xa0', ' ')
            text = re.sub(r'\s+', ' ', text)

            return text.strip()

        def normalize_header_value(value):
            text = cell_to_text(value).lower()

            text = text.replace('ё', 'е')
            text = text.replace('счёт', 'счет')
            text = text.replace('расчёт', 'расчет')

            text = re.sub(r'\s+', ' ', text)
            text = text.strip()
            text = text.strip(' .:;')

            return text

        alias_lookup = {}

        for field_name, aliases in HEADER_ALIASES.items():
            alias_lookup[field_name] = {
                normalize_header_value(alias)
                for alias in aliases
                if normalize_header_value(alias)
            }

        header_row_number = None
        header_values = None
        best_score = -1
        best_matches = set()

        max_scan_row = min(
            worksheet.max_row or 1,
            40
        )

        for row_number, row_values in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=max_scan_row,
                values_only=True
            ),
            start=1
        ):
            normalized_cells = {
                normalize_header_value(value)
                for value in row_values
                if normalize_header_value(value)
            }

            if not normalized_cells:
                continue

            matches = set()

            for field_name, aliases in alias_lookup.items():
                if normalized_cells.intersection(aliases):
                    matches.add(field_name)

            score = len(matches)

            if 'name' in matches:
                score += 10

            if 'inn' in matches:
                score += 4

            if 'kpp' in matches:
                score += 2

            if 'external_id_1c' in matches:
                score += 2

            if score > best_score:
                best_score = score
                best_matches = matches
                header_row_number = row_number
                header_values = row_values

        if not header_row_number or 'name' not in best_matches:
            raise Exception(
                'Не найдена строка заголовков XLSX. '
                'Ожидаются колонки вроде: Код, Контрагент, Наименование, ИНН, КПП.'
            )

        headers = []
        used_headers = {}

        for column_index, value in enumerate(
            header_values,
            start=1
        ):
            header = cell_to_text(value)

            if not header:
                headers.append(
                    f'__empty_column_{column_index}'
                )

                continue

            if header in used_headers:
                used_headers[header] += 1
                header = f'{header}__{used_headers[header]}'
            else:
                used_headers[header] = 1

            headers.append(header)

        rows = []

        for row_values in worksheet.iter_rows(
            min_row=header_row_number + 1,
            values_only=True
        ):
            row = {}
            has_data = False

            for header, value in zip(headers, row_values):
                if header.startswith('__empty_column_'):
                    continue

                row[header] = value

                if cell_to_text(value):
                    has_data = True

            if has_data:
                rows.append(row)

        return rows

    finally:
        try:
            workbook.close()
        except Exception:
            pass


def read_csv(file_path):

    encodings = [
        'utf-8-sig',
        'cp1251',
    ]

    for encoding in encodings:

        try:

            with open(
                file_path,
                'r',
                encoding=encoding,
                newline=''
            ) as file:

                sample = file.read(
                    4096
                )

                file.seek(
                    0
                )

                try:

                    dialect = csv.Sniffer().sniff(
                        sample,
                        delimiters=';,'
                    )

                except Exception:

                    dialect = csv.excel
                    dialect.delimiter = ';'

                reader = csv.DictReader(
                    file,
                    dialect=dialect
                )

                return list(
                    reader
                )

        except UnicodeDecodeError:

            continue

    raise Exception(
        'Не удалось определить кодировку CSV'
    )


def read_rows(file_path, sheet_name=None):

    file_path = Path(
        file_path
    )

    extension = file_path.suffix.lower()

    if extension in [
        '.xlsx',
        '.xlsm',
    ]:

        return read_xlsx(
            file_path,
            sheet_name=sheet_name
        )

    if extension == '.csv':

        return read_csv(
            file_path
        )

    raise Exception(
        'Поддерживаются только .xlsx, .xlsm, .csv'
    )


# ONE_C_DELETION_MARK_ACTIVE_FIX_V1-START
def normalize_1c_header_name(value):
    if value is None:
        return ''

    text = str(value)
    text = text.replace('\xa0', ' ')
    text = text.replace('ё', 'е')
    text = text.replace('счёт', 'счет')
    text = text.replace('расчёт', 'расчет')
    text = re.sub(r'\s+', ' ', text)
    text = text.strip().strip(' .:;').lower()

    return text


def normalize_1c_cell_text(value):
    if value is None:
        return ''

    text = str(value)
    text = text.replace('\xa0', ' ')
    text = re.sub(r'\s+', ' ', text)

    return text.strip()


def get_1c_row_value_by_headers(row, headers):
    normalized_headers = {
        normalize_1c_header_name(header)
        for header in headers
    }

    for key, value in row.items():
        if normalize_1c_header_name(key) in normalized_headers:
            return value

    return None


def parse_1c_bool(value, default=False):
    text = normalize_1c_cell_text(value).lower()
    text = text.replace('ё', 'е')

    if not text:
        return default

    if text in {
        '1',
        'true',
        'yes',
        'y',
        'да',
        'истина',
        'активен',
        'активный',
        'помечен',
    }:
        return True

    if text in {
        '0',
        'false',
        'no',
        'n',
        'нет',
        'ложь',
        'не активен',
        'неактивен',
        'не помечен',
    }:
        return False

    return default


def parse_counterparty_is_active(row):
    """
    В 1С колонка "Пометка удаления" означает обратное активности:
    - "Нет" => не помечен на удаление => активен;
    - "Да" => помечен на удаление => неактивен.
    """

    deletion_mark = get_1c_row_value_by_headers(
        row,
        {
            'Пометка удаления',
            'Deletion mark',
            'DeletionMark',
        }
    )

    if deletion_mark is not None and normalize_1c_cell_text(deletion_mark):
        return not parse_1c_bool(
            deletion_mark,
            default=False
        )

    active_value = get_1c_row_value_by_headers(
        row,
        {
            'Активен',
            'Активность',
            'Is active',
            'Active',
        }
    )

    if active_value is not None and normalize_1c_cell_text(active_value):
        return parse_1c_bool(
            active_value,
            default=True
        )

    return True
# ONE_C_DELETION_MARK_ACTIVE_FIX_V1-END


def import_counterparties_from_file(
    file_path,
    sheet_name=None,
    clear_ocr=False,
    deactivate_missing=False
):

    if clear_ocr:

        deleted_count, _ = Counterparty.objects.filter(
            source=Counterparty.SOURCE_OCR
        ).delete()

    else:

        deleted_count = 0

    rows = read_rows(
        file_path,
        sheet_name=sheet_name
    )

    created = 0
    updated = 0
    skipped = 0

    imported_external_ids = []

    for row in rows:

        external_id_1c = get_mapped_value(
            row,
            'external_id_1c'
        )

        name = get_mapped_value(
            row,
            'name'
        )

        full_name = get_mapped_value(
            row,
            'full_name'
        )

        inn = only_digits(
            get_mapped_value(
                row,
                'inn'
            )
        )

        kpp = only_digits(
            get_mapped_value(
                row,
                'kpp'
            )
        )

        bank_name = get_mapped_value(
            row,
            'bank_name'
        )

        bik = only_digits(
            get_mapped_value(
                row,
                'bik'
            )
        )

        account_number = only_digits(
            get_mapped_value(
                row,
                'account_number'
            )
        )

        correspondent_account = only_digits(
            get_mapped_value(
                row,
                'correspondent_account'
            )
        )

        is_active = parse_bool(
            get_mapped_value(
                row,
                'is_active'
            ),
            default=True
        )

        if parse_deletion_mark(
            get_mapped_value(
                row,
                'deletion_mark'
            )
        ):

            is_active = False

        if not name and full_name:

            name = full_name

        if not name:

            skipped += 1

            continue

        counterparty = None

        if external_id_1c:

            imported_external_ids.append(
                external_id_1c
            )

            counterparty = Counterparty.objects.filter(
                external_id_1c=external_id_1c
            ).first()

        if not counterparty and inn and kpp:

            counterparty = Counterparty.objects.filter(
                inn=inn,
                kpp=kpp,
                source=Counterparty.SOURCE_1C
            ).first()

        if not counterparty and inn:

            counterparty = Counterparty.objects.filter(
                inn=inn,
                source=Counterparty.SOURCE_1C
            ).first()

        if not counterparty:

            counterparty = Counterparty.objects.filter(
                name__iexact=name,
                source=Counterparty.SOURCE_1C
            ).first()

        if counterparty:

            was_created = False

        else:

            counterparty = Counterparty()

            was_created = True

        counterparty.external_id_1c = (
            external_id_1c
            or counterparty.external_id_1c
        )

        counterparty.name = name

        counterparty.full_name = (
            full_name
            or counterparty.full_name
        )

        counterparty.inn = (
            inn
            or counterparty.inn
        )

        counterparty.kpp = (
            kpp
            or counterparty.kpp
        )

        counterparty.bank_name = (
            bank_name
            or counterparty.bank_name
        )

        counterparty.bik = (
            bik
            or counterparty.bik
        )

        counterparty.account_number = (
            account_number
            or counterparty.account_number
        )

        counterparty.correspondent_account = (
            correspondent_account
            or counterparty.correspondent_account
        )

        counterparty.source = Counterparty.SOURCE_1C

        counterparty.is_active = parse_counterparty_is_active(row)

        counterparty.synced_at = timezone.now()

        counterparty.sync_comment = (
            'Импортировано из 1С через веб-интерфейс'
        )

        counterparty.save()

        if was_created:

            created += 1

        else:

            updated += 1

    deactivated = 0

    if deactivate_missing and imported_external_ids:

        deactivated = Counterparty.objects.filter(
            source=Counterparty.SOURCE_1C
        ).exclude(
            external_id_1c__in=imported_external_ids
        ).update(
            is_active=False
        )

    return {
        'rows': len(rows),
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'deleted_ocr': deleted_count,
        'deactivated': deactivated,
        'total_1c': Counterparty.objects.filter(
            source=Counterparty.SOURCE_1C
        ).count(),
    }

# ONE_C_HEADER_ALIASES_1C_EXPORT_V1-START
# Дополнительные варианты заголовков из стандартной выгрузки 1С.
_extra_1c_header_aliases = {
    'external_id_1c': [
        'код',
        'код 1с',
    ],
    'name': [
        'контрагент',
        'наименование',
        'краткое наименование',
    ],
    'full_name': [
        'полное наименование',
    ],
    'inn': [
        'инн',
    ],
    'kpp': [
        'кпп',
    ],
    'bank_name': [
        'банк',
        'наименование банка',
        'банковский счет.банк.наименование',
        'банковский счёт.банк.наименование',
    ],
    'bik': [
        'бик',
        'банковский счет.банк.бик',
        'банковский счёт.банк.бик',
    ],
    'account_number': [
        'номер счета',
        'номер счёта',
        'расчетный счет',
        'расчётный счёт',
        'банковский счет.номер счета',
        'банковский счёт.номер счёта',
    ],
    'correspondent_account': [
        'корр. счет',
        'корр. счёт',
        'корреспондентский счет',
        'корреспондентский счёт',
        'банковский счет.банк.корр. счет',
        'банковский счёт.банк.корр. счёт',
    ],
    'is_active': [
        'пометка удаления',
    ],
}

for _field_name, _aliases in _extra_1c_header_aliases.items():
    _target_aliases = HEADER_ALIASES.setdefault(_field_name, [])

    for _alias in _aliases:
        if _alias not in _target_aliases:
            _target_aliases.append(_alias)
# ONE_C_HEADER_ALIASES_1C_EXPORT_V1-END
